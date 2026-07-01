# tools/portfolio.py
"""Portfolio runner: run the engine once per property and roll up a summary.

Finance-readable summary
------------------------
This is the one button that runs every property in the portfolio and collects a
single side-by-side summary. It reads portfolio.yaml (the list of properties and
their on/off switches), runs the same ``python -m src.engine`` per property that
a person would run by hand, and gathers one headline row per property into a CSV
and a formatted Excel workbook.

Phase 6 / S5 note: the runner now handles a mixed portfolio cleanly. A property
with a mortgage runs exactly as before (it passes ``--actuals`` and reads the
monthly schedule for KPIs). A property with no mortgage (owned-outright,
declared in portfolio.yaml without an ``actuals`` line) runs through the
engine's valuation-only path: ``--actuals`` is omitted and the summary row is
built from ``valuation_schedule.csv`` instead of the loan schedule. Per-property
enable already worked via the ``enabled`` flag and is unchanged.

Phase 8 / S2 note: the local ``slugify`` was removed and is now imported from
``src.engine.helpers``. The output folder slug and the per-property workbook
name (``<slug>_model.xlsx``) are produced by that one function, so the file and
its folder can never disagree.

Phase 8 / S3 note: the per-property CSVs the engine writes now live under each
property's ``output.csv_subdir`` sub-folder (default ``csv``), so the runner
loads each property's inputs once and reads ``schedule_monthly.csv``,
``events_daily.csv`` and ``valuation_schedule.csv`` from that csv/ folder. The
rolled-up ``portfolio_summary.csv`` is likewise written under a top-level
``csv/`` folder; the ``portfolio_summary.xlsx`` workbook stays at the output
root.

Phase 8 / S4 note: the rollup itself is rebuilt. Earlier versions listed several
KPI keys in a prefer-order that ``compute_baseline_kpis`` never returned, so
those columns silently never appeared. This version builds each row explicitly
from the monthly schedule and reads the tax-year file, so every promised column
is populated and the column order is locked (see ``LOCKED_SUMMARY_COLUMNS``).
The live-position figures are taken from the monthly row at the as-of date
(derived the same way the engine CLI derives it, so the rollup ties out to each
property's own Summary sheet), because the schedule projects all the way to
payoff and its final row is the payoff month rather than today. Engine maths is
unchanged; this is a read-and-aggregate layer only.

Phase 8 / S5 note: the snapshot date is now made explicit. ``as_of_date`` is
added as the first rollup column so a reader can see, on the rollup itself, the
date the whole row is a snapshot of. It was deliberately left out of the locked
15 in S4; surfacing it is the only column change in S5 and moves the lock from
15 to 16 columns. For a mortgage row the value is the same deterministic as-of
date the live-position figures are already anchored on; for a valuation-only row
it is the run date, matching the as-of-today value that row already reports.
Engine maths is still untouched.

Technical summary
-----------------
``run_engine_cli`` appends ``--actuals`` only when one is given. The main loop
classifies each enabled property as mortgage-bearing or valuation-only, runs the
right engine path, and appends the matching summary row. ``_mortgage_summary_row``
builds the locked row from the monthly schedule, the event log (payoff date via
``compute_baseline_kpis``), and the optional tax-year file; ``_valuation_summary_row``
builds the no-loan row. The combined DataFrame is reindexed onto the locked
column order, so a valuation-only property's loan columns are simply blank.
"""
from __future__ import annotations
import argparse, subprocess, sys
import datetime as _dt
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd
import yaml
from openpyxl.styles import Font  # Phase 6 / S6: build bold fonts directly (Font.copy() is deprecated)

from src.engine import load_inputs  # to pass real inputs to KPIs if supported
from src.metrics import compute_baseline_kpis
# Phase 8 / S2: one canonical slugify lives in the engine so the workbook slug
# and this output-folder slug are produced by the same function and cannot drift.
from src.engine.helpers import slugify
# Phase 2 path resolver: output root and per-property paths come from the config
# layer instead of being assumed relative to the current working directory.
from src.paths import resolve_out_dir, resolve_relative

# Phase 6 / S5: kinds that carry no mortgage and therefore run the valuation-only
# path. Mirrors the canonical owned-outright spellings the schema accepts so the
# runner agrees with the engine without importing schema internals.
_VALUATION_ONLY_KINDS = {"owned_outright", "owned-outright", "outright", "owned"}

# Phase 8 / S3: default CSV sub-folder for the top-level portfolio rollup.
DEFAULT_CSV_SUBDIR = "csv"

# Phase 8 / S4: the locked final column order for the rebuilt rollup. The
# friendly labels from the session scope map one-to-one onto these machine-
# friendly keys (kept snake_case to match every other CSV the suite pins):
#   as_of_date                   -> As-of date (Phase 8 / S5)
#   property_name                -> Property
#   property_kind                -> Kind
#   tax_enabled                  -> Tax on
#   current_balance              -> Current balance
#   property_value               -> Property value
#   ltv                          -> LTV
#   current_annual_rate          -> Current rate
#   contractual_payment          -> Contractual payment
#   current_overpayment          -> Current overpayment per month
#   total_overpaid_to_date       -> Total overpaid to date
#   total_difference             -> Difference
#   overpayment_mismatch_months  -> Mismatch months
#   payoff_date                  -> Projected payoff date
#   current_year_interest        -> Annual interest (current year)
#   tax_deductible_interest      -> Tax-deductible interest (when tax on)
#
# Phase 8 / S5: as_of_date leads the list so the snapshot date is visible on the
# rollup itself. It was kept out of the locked 15 in S4; making it explicit is
# the only column change in S5 and moves the lock from 15 to 16 columns.
LOCKED_SUMMARY_COLUMNS = [
    "as_of_date",
    "property_name", "property_kind", "tax_enabled",
    "current_balance", "property_value", "ltv", "current_annual_rate",
    "contractual_payment", "current_overpayment", "total_overpaid_to_date",
    "total_difference", "overpayment_mismatch_months",
    "payoff_date", "current_year_interest", "tax_deductible_interest",
]

def load_portfolio(p: Path) -> Dict:
    """Read portfolio.yaml into a dict and check it carries a properties list.

    Finance note: portfolio.yaml is the master list of which properties exist
    and which are switched on. A missing 'properties' list is a hard error
    because there would be nothing to run.
    """
    raw = yaml.safe_load(p.read_text())
    assert "properties" in raw and isinstance(raw["properties"], list), "portfolio.yaml missing 'properties' list"
    return raw

def run_engine_cli(inputs_path: Path, actuals_path: Optional[Path], out_dir: Path) -> None:
    """Run ``python -m src.engine`` once for a single property.

    Finance note: this shells out to exactly the command a person would type by
    hand, so the portfolio runner and a manual run produce identical per-property
    files. A valuation-only property has no bank loan to reconcile, so
    ``actuals_path`` is None and ``--actuals`` is left off; the engine then takes
    its no-mortgage valuation-only path.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    # Build the command incrementally: --actuals is only added for a mortgage
    # property. Omitting it is what routes an owned-outright property to the
    # engine's valuation-only path.
    cmd = [sys.executable, "-m", "src.engine", "--inputs", str(inputs_path)]
    if actuals_path is not None:
        cmd += ["--actuals", str(actuals_path)]
    cmd += ["--out", str(out_dir)]
    subprocess.run(cmd, check=True)

# ---- Shared coercion helpers -------------------------------------------------

def _to_date(value) -> Optional[_dt.date]:
    """Coerce a date-like value (date, Timestamp, or ISO string) to a date.

    Returns None for anything that cannot be read as a date, so every caller can
    treat a missing or malformed date as 'unknown' rather than crashing.
    """
    if value is None:
        return None
    # pandas Timestamp is a subclass of datetime, so this catches both.
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.date()

def _count_true(series: pd.Series) -> int:
    """Count truthy flags in a column that may be bool, numeric, or text.

    pandas usually reads the overpayment_mismatch column back as proper bools,
    but a CSV round-trip can also surface 'True'/'False' strings, so this coerces
    both shapes to a single integer count of the months flagged.
    """
    filled = series.fillna(False)
    if filled.dtype == bool:
        return int(filled.sum())
    text = filled.astype(str).str.strip().str.lower()
    return int(text.isin(["true", "1", "yes"]).sum())

def _derive_as_of(csv_dir: Path, inputs_path: Path) -> Optional[_dt.date]:
    """Return the deterministic 'as-of' date the current snapshot is taken at.

    Finance note: 'now' for a mortgage is the latest point we have real bank
    data for. This mirrors the engine CLI exactly so the rollup ties out to each
    property's own Summary sheet: start from the newest reconciled bank actual
    (the latest bank_date in reconcile.csv that the model lined up against), then
    prefer a newer portal snapshot date from the inputs' reconcile.snapshots
    block when one exists. Returns None when neither source is available, and the
    caller then falls back to the final monthly row.
    """
    as_of: Optional[_dt.date] = None
    # 1) Latest reconciled bank actual from reconcile.csv.
    reconcile_csv = csv_dir / "reconcile.csv"
    if reconcile_csv.exists():
        rec = pd.read_csv(reconcile_csv, parse_dates=["bank_date"])
        # Only rows the model reconciled against count as a real bank actual.
        if "model_balance" in rec.columns:
            rec = rec.dropna(subset=["model_balance"])
        if not rec.empty and "bank_date" in rec.columns:
            as_of = _to_date(rec["bank_date"].max())
    # 2) Prefer a newer portal snapshot if the inputs declare one. The snapshots
    #    live in the raw YAML (the typed Inputs object does not carry them), so
    #    read them the same way the engine CLI does.
    try:
        raw = yaml.safe_load(Path(inputs_path).read_text())
    except Exception:
        raw = {}
    snaps = ((raw or {}).get("reconcile") or {}).get("snapshots") or []
    snap_dates = [d for d in (_to_date(s.get("date")) for s in snaps) if d is not None]
    if snap_dates:
        latest_snap = max(snap_dates)
        if as_of is None or latest_snap > as_of:
            as_of = latest_snap
    return as_of

def _current_row(monthly: pd.DataFrame, as_of: Optional[_dt.date]) -> pd.Series:
    """Return the last monthly row on or before the as-of date (the live position).

    Finance note: the schedule projects all the way to payoff, so its final row
    is the payoff month (a zero balance), not 'now'. The current snapshot is the
    last row whose month falls on or before the as-of date. When the as-of date
    is unknown, fall back to the final row so the builder still returns a value.
    """
    if as_of is not None and "month_start" in monthly.columns:
        month_starts = monthly["month_start"].apply(_to_date)
        mask = month_starts.apply(lambda d: d is not None and d <= as_of)
        if mask.any():
            return monthly.loc[mask].iloc[-1]
    return monthly.iloc[-1]

def _tax_deductible_for_year(csv_dir: Path, year: Optional[int]) -> Optional[float]:
    """Return the Section 97 allowable interest for a year from tax_year.csv.

    Finance note: this is the only cross-file read in the rollup. It opens the
    property's tax_year.csv (written only when rental tax is on) and returns the
    allowable_interest_s97 figure for the given calendar year. A missing file,
    missing column, or absent year returns None so the column blanks gracefully
    rather than guessing.
    """
    if year is None:
        return None
    tax_csv = csv_dir / "tax_year.csv"
    if not tax_csv.exists():
        return None
    tax = pd.read_csv(tax_csv)
    if "year" not in tax.columns or "allowable_interest_s97" not in tax.columns:
        return None
    hit = tax.loc[tax["year"] == year]
    if hit.empty:
        return None
    value = hit.iloc[0]["allowable_interest_s97"]
    return float(value) if pd.notna(value) else None

# ---- XLSX formatting helpers (lightweight, values-only) ----------------------

def _header_map(ws):
    """Map each column header text to its 1-based column index."""
    return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

def fmt_money(ws, col_name):
    """Apply a euro money format to a named column, if it is present."""
    col = _header_map(ws).get(col_name)
    if not col: return
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=col).number_format = "\u20ac#,##0.00"

def fmt_pct(ws, col_name):
    """Apply a percent format to a named column, if it is present."""
    col = _header_map(ws).get(col_name)
    if not col: return
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=col).number_format = "0.00%"

def fmt_date(ws, col_name):
    """Apply an ISO date format to a named column, if it is present."""
    col = _header_map(ws).get(col_name)
    if not col: return
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=col).number_format = "yyyy-mm-dd"

def write_summary_xlsx(df: pd.DataFrame, path: Path):
    """Write the portfolio summary DataFrame to a formatted Excel workbook.

    Finance note: this is the one-look portfolio sheet. It bolds the header,
    adds a filter and table stripes, and applies money / percent / date formats
    so the rolled-up numbers read cleanly.
    """
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        df.to_excel(xl, index=False, sheet_name="Portfolio")
        ws = xl.sheets["Portfolio"]

        # bold header + autofilter + widths
        for cell in ws[1]:
            # Phase 6 / S6: openpyxl 3.x deprecated Font.copy(); build a new bold
            # Font instead. Header cells start from the default font, so a plain
            # bold Font reproduces the previous styling exactly.
            cell.font = Font(bold=True)
        ws.auto_filter.ref = ws.dimensions
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 24

        # Table stripes if available
        try:
            from openpyxl.worksheet.table import Table, TableStyleInfo
            ref = f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}{ws.max_row}"
            tbl = Table(displayName="TblPortfolio", ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tbl)
        except Exception:
            pass

        # Phase 8 / S4: number formats follow the rebuilt locked columns. Money
        # columns are the euro figures; the two rates are percentages; the date
        # columns are dates. The mismatch-month count is a plain integer and
        # needs no mask.
        # Phase 8 / S5: as_of_date joins payoff_date in the date-formatted set.
        money_cols = [
            "current_balance", "property_value", "contractual_payment",
            "current_overpayment", "total_overpaid_to_date", "total_difference",
            "current_year_interest", "tax_deductible_interest",
        ]
        pct_cols = ["ltv", "current_annual_rate"]
        date_cols = ["as_of_date", "payoff_date"]

        for c in money_cols: fmt_money(ws, c)
        for c in pct_cols: fmt_pct(ws, c)
        for c in date_cols: fmt_date(ws, c)

# ---- Summary-row helpers -----------------------------------------------------

def _is_valuation_only(p: Dict) -> bool:
    """Decide whether a portfolio entry runs the no-mortgage valuation-only path.

    Finance note: an owned-outright property has no bank loan, so it declares no
    ``actuals`` file and the engine only tracks its value. A missing ``actuals``
    line is the primary signal (it is what makes the engine omit the loan path);
    an explicit owned-outright kind is also accepted for clarity.
    """
    if not p.get("actuals"):
        return True
    return str(p.get("property_kind", "")).strip().lower() in _VALUATION_ONLY_KINDS

def _mortgage_summary_row(
    monthly: pd.DataFrame,
    events: pd.DataFrame,
    prop_inputs,
    csv_dir: Path,
    inputs_path: Path,
    name: str,
    kind: str,
    tax_enabled: bool,
) -> Dict:
    """Build one locked portfolio-summary row for a mortgage-bearing property.

    Finance note: this is the rebuilt rollup row. Every promised column is
    populated and sourced explicitly, so nothing is silently dropped:
      * As-of date (Phase 8 / S5) is the snapshot date the rest of the row is
        taken at, surfaced as the first column so the date is visible.
      * Current balance / Property value / LTV / Current rate / Contractual
        payment / Current overpayment per month come from the current snapshot
        row, which is the last monthly row on or before the as-of date. The
        schedule runs to payoff, so its final row is the payoff month, not
        'now'; anchoring on the as-of date is what makes these read as the live
        position.
      * Total overpaid to date sums the agreed overpayment actually made up to
        the as-of date (future projected overpayments are excluded).
      * Difference and Mismatch months are the Phase 7 attribution health: the
        total unattributed Difference across the schedule and the count of
        months the engine flagged as an overpayment mismatch.
      * Annual interest (current year) sums the modelled interest posted in the
        as-of calendar year. Tax-deductible interest is the Section 97 allowable
        interest for that year from tax_year.csv, populated only when rental tax
        is on and the file exists.
      * Projected payoff date is the engine's payoff date (the first month the
        model balance clears), read via compute_baseline_kpis from the events.
    """
    # The deterministic as-of date and the matching current snapshot row.
    as_of = _derive_as_of(csv_dir, inputs_path)
    current = _current_row(monthly, as_of)

    # The calendar year the 'current year' figures belong to: the as-of year
    # when known, otherwise the year of the snapshot row we fell back to.
    current_month = _to_date(current.get("month_start"))
    current_year = as_of.year if as_of is not None else (current_month.year if current_month else None)

    # Reuse the shared KPI helper purely for the payoff date so the rollup and
    # the per-property outputs agree on when the loan clears. The try/except
    # keeps the older one-argument signature working.
    try:
        kpis = compute_baseline_kpis(prop_inputs, monthly, events)
    except TypeError:
        kpis = compute_baseline_kpis(monthly)
    payoff_date = kpis.get("payoff_date")

    # Total overpaid to date: agreed overpayment actually made up to the as-of
    # date. Without an as-of date, fall back to the whole-schedule total.
    if as_of is not None and "month_start" in monthly.columns:
        month_starts = monthly["month_start"].apply(_to_date)
        to_date_mask = month_starts.apply(lambda d: d is not None and d <= as_of)
        total_overpaid_to_date = float(monthly.loc[to_date_mask, "overpayment"].fillna(0.0).sum())
    elif "overpayment" in monthly.columns:
        total_overpaid_to_date = float(monthly["overpayment"].fillna(0.0).sum())
    else:
        total_overpaid_to_date = None

    # Annual interest for the current calendar year, summed from the modelled
    # interest the engine posted in that year.
    current_year_interest = None
    if current_year is not None and "posting_year" in monthly.columns and "interest_used" in monthly.columns:
        year_mask = monthly["posting_year"] == current_year
        current_year_interest = float(monthly.loc[year_mask, "interest_used"].fillna(0.0).sum())

    # Attribution health: the total unattributed Difference and the number of
    # months flagged as an overpayment mismatch.
    total_difference = float(monthly["difference"].fillna(0.0).sum()) if "difference" in monthly.columns else None
    mismatch_months = _count_true(monthly["overpayment_mismatch"]) if "overpayment_mismatch" in monthly.columns else None

    # Tax-deductible interest for the current year (Section 97 allowable), read
    # from tax_year.csv only when rental tax is on and the file exists.
    tax_deductible_interest = _tax_deductible_for_year(csv_dir, current_year) if tax_enabled else None

    def _cell(col: str) -> Optional[float]:
        """Read a numeric cell from the current snapshot row as a float or None."""
        if col not in current.index:
            return None
        value = current[col]
        return float(value) if pd.notna(value) else None

    return {
        # Phase 8 / S5: lead with the snapshot date the whole row is taken at.
        "as_of_date": as_of,
        "property_name": name,
        "property_kind": kind,
        "tax_enabled": tax_enabled,
        "current_balance": _cell("model_eom_balance"),
        "property_value": _cell("property_value"),
        "ltv": _cell("ltv_model_eom"),
        "current_annual_rate": _cell("annual_rate"),
        "contractual_payment": _cell("contractual"),
        "current_overpayment": _cell("overpayment"),
        "total_overpaid_to_date": total_overpaid_to_date,
        "total_difference": total_difference,
        "overpayment_mismatch_months": mismatch_months,
        "payoff_date": _to_date(payoff_date),
        "current_year_interest": current_year_interest,
        "tax_deductible_interest": tax_deductible_interest,
    }

def _valuation_summary_row(csv_dir: Path, name: str, kind: str, tax_enabled: bool) -> Dict:
    """Build one locked portfolio-summary row for a valuation-only property.

    Finance note: an owned-outright property has no loan, so every loan and
    attribution column is blank. Its one meaningful figure in the rollup is the
    current property value, read from valuation_schedule.csv at the row on or
    before today. The as-of date here is the run date, because a no-loan
    property has no bank actuals to date it from, and the value it reports is the
    as-of-today value; every other locked column is left unset and shows blank
    in the combined summary.

    Phase 8 / S3: ``csv_dir`` is the property's csv/ sub-folder, where the engine
    now writes the CSV.
    """
    val_csv = csv_dir / "valuation_schedule.csv"
    if not val_csv.exists():
        raise FileNotFoundError(f"Expected valuation CSV missing: {val_csv}")
    sched = pd.read_csv(val_csv, parse_dates=["month_start"])
    # Current value: the last row on or before today, falling back to the final
    # modelled row when the series begins in the future.
    today = _dt.date.today()
    month_starts = sched["month_start"].apply(_to_date)
    mask = month_starts.apply(lambda d: d is not None and d <= today)
    current = sched.loc[mask].iloc[-1] if mask.any() else sched.iloc[-1]
    return {
        # Phase 8 / S5: the snapshot date for a no-loan property is the run date,
        # matching the as-of-today property value reported below.
        "as_of_date": today,
        "property_name": name,
        "property_kind": kind,
        "tax_enabled": tax_enabled,
        "property_value": float(current["property_value"]),
    }

# ---- Main --------------------------------------------------------------------

def main():
    """Run every enabled property and write the rolled-up portfolio summary.

    Finance note: reads portfolio.yaml, runs each switched-on property through
    the engine (a mortgage property with its bank actuals, an owned-outright
    property through the valuation-only path), and gathers one headline row per
    property into a CSV and a formatted Excel workbook.
    """
    ap = argparse.ArgumentParser(description="Portfolio runner (delegates to engine CLI per property)")
    ap.add_argument("--portfolio", type=Path, required=True, help="Path to data/portfolio.yaml")
    # --out is optional now.  When omitted, the output root is resolved through
    # the config layer (CLI > MORTGAGE_OUT_DIR > paths.local.yaml > <repo>/out).
    ap.add_argument("--out", type=Path, default=None, help="Root output folder (overrides config)")
    ap.add_argument("--only", type=str, default=None, help="Run only this property name (exact match)")
    args = ap.parse_args()

    port = load_portfolio(args.portfolio)
    props = port["properties"]
    if args.only:
        props = [p for p in props if str(p.get("name", "")) == args.only]

    # Resolve the output root through the config layer.  Passing the raw CLI value
    # (or None) keeps an explicit --out as the highest-priority source.
    out_root = resolve_out_dir(str(args.out) if args.out is not None else None)
    out_root.mkdir(parents=True, exist_ok=True)

    rows: List[Dict] = []
    for p in props:
        if not p.get("enabled", False):
            continue

        name = str(p["name"])
        kind = str(p.get("property_kind", ""))
        tax_enabled = bool(p.get("tax_enabled", False))
        # Resolve per-property paths relative to the portfolio.yaml location so
        # relative entries do not depend on the current working directory.
        inputs_path = resolve_relative(args.portfolio, p["inputs"])
        slug = p.get("out_dir") or slugify(name)
        out_dir = out_root / slug

        # Phase 8 / S3: the engine writes every CSV under a per-property csv/
        # sub-folder, so the runner reads from the same place. Load the
        # property's own inputs once to read output.csv_subdir (default "csv";
        # an empty value means the old flat layout); the object is reused for the
        # KPI call below.
        prop_inputs = load_inputs(inputs_path)
        csv_subdir = prop_inputs.output.csv_subdir
        csv_dir = (out_dir / csv_subdir) if csv_subdir else out_dir

        # Phase 6 / S5: a no-mortgage property runs the valuation-only path. It
        # has no bank actuals and emits valuation_schedule.csv rather than the
        # loan schedule, so it gets its own run + summary branch.
        if _is_valuation_only(p):
            run_engine_cli(inputs_path, None, out_dir)
            rows.append(_valuation_summary_row(csv_dir, name, kind, tax_enabled))
            continue

        # Mortgage property: pass the bank actuals and read the monthly schedule.
        actuals_path = resolve_relative(args.portfolio, p["actuals"])

        # 1) Run engine CLI for FULL outputs (XLSX + CSVs [+ tax if enabled]).
        run_engine_cli(inputs_path, actuals_path, out_dir)

        # 2) Read the per-property CSVs the row is built from.
        monthly_csv = csv_dir / "schedule_monthly.csv"
        events_csv = csv_dir / "events_daily.csv"
        if not monthly_csv.exists():
            raise FileNotFoundError(f"Expected monthly CSV missing: {monthly_csv}")
        monthly = pd.read_csv(
            monthly_csv,
            parse_dates=["month_start", "payment_date", "posting_date"],
        )
        events = pd.read_csv(events_csv, parse_dates=["date"]) if events_csv.exists() else pd.DataFrame()

        # 3) Build the locked rollup row for this property.
        rows.append(
            _mortgage_summary_row(
                monthly, events, prop_inputs, csv_dir, inputs_path, name, kind, tax_enabled
            )
        )

    # 4) Write portfolio summary (CSV + nicely formatted XLSX).
    if rows:
        df = pd.DataFrame(rows)
        # Lock the final column order exactly. reindex adds any column a row did
        # not supply (a valuation-only property omits the loan columns) as blank,
        # and drops nothing because every produced key is in the locked list.
        df = df.reindex(columns=LOCKED_SUMMARY_COLUMNS)

        # Phase 8 / S3: the rollup CSV is demoted into a top-level csv/ folder to
        # match the per-property layout. The portfolio knob is read from an
        # optional top-level output.csv_subdir in portfolio.yaml, defaulting to
        # "csv"; an empty value keeps the rollup at the output root. The summary
        # workbook stays at the output root.
        raw_rollup_subdir = (port.get("output") or {}).get("csv_subdir", DEFAULT_CSV_SUBDIR)
        rollup_subdir = ("" if raw_rollup_subdir is None else str(raw_rollup_subdir)).strip().strip("/\\")
        rollup_csv_dir = (out_root / rollup_subdir) if rollup_subdir else out_root
        rollup_csv_dir.mkdir(parents=True, exist_ok=True)
        df.to_csv(rollup_csv_dir / "portfolio_summary.csv", index=False)
        write_summary_xlsx(df, out_root / "portfolio_summary.xlsx")

    print(f"Wrote portfolio outputs under: {out_root.resolve()}")

if __name__ == "__main__":
    main()