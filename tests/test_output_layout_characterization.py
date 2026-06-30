"""Characterization test that pins the v1.8.0 output layout on disk.

Finance-readable summary
------------------------
Before Phase 8 reshapes where the report files live and what they are called,
this test takes a precise photograph of what the engine produces today. It runs
the model against the three bundled sample properties, one of each kind, and
records the exact file names, the worksheet tabs inside each Excel workbook, and
the column headings of every CSV. If a later change adds, drops, renames, or
reorders any of those, this test fails and names what moved, so the restructure
in S2 to S4 is always provably intentional and never an accident.

The three sample properties deliberately cover the whole output matrix:
  * Property A is an investment let: it takes the mortgage path and, because
    rental tax is on, it emits the full tax sheets and tax CSVs.
  * Property B is a primary residence: it takes the same mortgage path but with
    tax switched off, so the tax sheets and tax CSVs are absent.
  * Property C is owned outright with no loan: it takes the valuation-only path
    and emits a value-over-time workbook and CSV instead.
The portfolio rollup (one row per enabled property) is pinned alongside them.

Technical summary
-----------------
Runs ``python -m src.engine`` once per sample property into its own temp folder
(A and B pass ``--actuals`` and take the mortgage path; C omits ``--actuals``
and the engine branches to the valuation-only path on ``meta.mortgage_enabled``
being False), and runs ``tools.baseline`` then ``tools.portfolio`` once for the
rollup, mirroring the golden-master invocation. It then asserts the produced
file set, the openpyxl sheet-name order per workbook, and the pandas-read column
order per CSV against constants captured from the locked v1.8.0 golden fixtures
and the two writer paths. This is a shape test only; it asserts no numbers, so
it never overlaps with the golden-master numeric contract.

Phase 8 / S3 note
-----------------
The CSV outputs now live one level below each output root, in a ``csv/``
subfolder (the ``output.csv_subdir`` knob, default ``csv``); only the workbooks
stay at the root. This test therefore looks for each workbook at the property
root and every CSV inside ``csv/``, and the rollup CSV inside a top-level
``csv/``. The committed golden fixtures are intentionally left where they are;
the S5 re-baseline owns moving them.
"""
from __future__ import annotations

import os
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import pytest
from openpyxl import load_workbook


def _find_repo_root(start: Path) -> Path:
    """Return the repo root by walking up from this file.

    The repo root is the first ancestor that holds both the ``tools`` package
    and the ``data_sample`` directory side by side, so the test resolves
    correctly no matter how deep under the repo the file is placed. This mirrors
    the golden-master test's resolver so the two suites agree on the root.
    """
    # Check this file's own directory first, then each parent in turn.
    for candidate in (start, *start.parents):
        # Only the repo root holds both of these together.
        if (candidate / "tools").is_dir() and (candidate / "data_sample").is_dir():
            return candidate
    # Fail loudly rather than silently resolving to the wrong directory.
    raise RuntimeError(
        "Could not locate the mortflow repo root: no ancestor of "
        f"{start} contains both tools/ and data_sample/."
    )


# Resolve the repo root from this file's location, robust to where it sits.
REPO_ROOT = _find_repo_root(Path(__file__).resolve().parent)
# The bundled de-identified sample data the whole suite runs against.
DATA_SAMPLE = REPO_ROOT / "data_sample"
# The sample portfolio file the rollup runner reads (Property A is the only
# enabled entry, so the rollup pins a single mortgage row).
SAMPLE_PORTFOLIO = DATA_SAMPLE / "portfolio.yaml"

# Phase 8 / S3: every CSV is now written one level below the property (and the
# portfolio) output root, into this subfolder. It mirrors the default of the
# output.csv_subdir knob ("csv"); only the workbooks stay at the root. The
# committed golden fixtures are unchanged and the S5 re-baseline owns moving them.
CSV_SUBDIR = "csv"


# ---------------------------------------------------------------------------
# Locked v1.8.0 column vocabularies.
#
# These lists are the contract this test pins. They were captured from the
# committed golden fixtures (the authoritative record of a verified run) and the
# two writer paths. The monthly, reconcile, and events column sets are shared by
# every mortgage property because the engine builds them from the loan schedule
# regardless of whether rental tax is switched on.
# ---------------------------------------------------------------------------

# Monthly schedule. Carries the Phase 7 attribution vocabulary: contractual,
# overpayment, lump, total_paid, and difference (the unattributed remainder).
MONTHLY_COLUMNS: List[str] = [
    "ym", "month_start", "payment_date",
    "contractual", "overpayment", "lump", "total_paid", "difference",
    "overpayment_mismatch",
    "interest_used", "principal_paid", "annual_rate",
    "bank_posted_interest_present", "posting_date", "posting_year",
    "model_eom_balance", "bank_eom_running_balance", "eom_diff_model_minus_bank",
    "property_value", "ltv_model_eom", "ltv_bank_eom",
]

# Bank reconciliation: each modelled month lined up against a bank posting, with
# the tolerance verdict columns the reconciler appends.
RECONCILE_COLUMNS: List[str] = [
    "bank_date", "type", "amount", "bank_running_balance", "ym",
    "model_amount", "model_balance", "diff_model_minus_bank",
    "ok_within_1c", "ok_within_abs_eur", "ok_label", "ok_reason",
]

# Daily event log: one row per cashflow event, with the running balance and LTV.
EVENTS_DAILY_COLUMNS: List[str] = [
    "date", "kind", "amount", "balance", "property_value", "ltv_after_event", "ym",
]

# Tax-year roll-up (Irish Section 97 allowable interest), one row per tax year.
TAX_YEAR_COLUMNS: List[str] = [
    "year", "interest_posted", "allowable_interest_s97", "principal_paid",
    "avg_occupancy_ratio", "deductible_pct",
]

# Monthly tax audit trail: the per-month working behind the tax-year figures.
TAX_AUDIT_COLUMNS: List[str] = [
    "ym", "month_start", "posting_date", "posting_year",
    "interest_used", "principal_paid",
    "days_let_or_available", "days_in_month",
    "occupancy_ratio", "deductible_pct", "allowable_interest_s97",
]

# Valuation-only value-over-time series for a no-loan property.
VALUATION_SCHEDULE_COLUMNS: List[str] = ["month_start", "ym", "property_value"]

# Portfolio rollup. Phase 8 / S4 rebuilt this list. The runner no longer carries
# the dead KPI columns it used to list but never emit (property_value_asof,
# ltv_asof, current_annual_rate, principal_excl_unposted); instead it sources
# every promised column from a real place: the live position from the current
# snapshot row of the monthly schedule, the agreed overpayment to date, the
# Phase 7 attribution health (total_difference plus a count of
# overpayment_mismatch months), the current-year interest, the projected payoff
# date, and the Section 97 tax-deductible interest. This pins the rebuilt
# 15-column contract, in order. (The golden-master fixture re-baseline that
# follows from this rebuild is owned by S5.)
PORTFOLIO_SUMMARY_COLUMNS: List[str] = [
    "property_name", "property_kind", "tax_enabled",
    "current_balance", "property_value", "ltv", "current_annual_rate",
    "contractual_payment", "current_overpayment", "total_overpaid_to_date",
    "total_difference", "overpayment_mismatch_months",
    "payoff_date", "current_year_interest", "tax_deductible_interest",
]


@dataclass(frozen=True)
class PropertyShape:
    """The locked output shape for one sample property.

    Finance note: groups everything we expect a single property to produce, so
    one place names its kind, its workbook, the worksheet tabs in that workbook,
    and the column headings of each CSV it writes.
    """

    slug: str                      # output sub-folder name, e.g. "property-a"
    kind_label: str                # plain-language kind, for readable failures
    inputs: Path                   # the property's sample inputs YAML
    actuals: Optional[Path]        # bank actuals CSV, or None for valuation-only
    workbook: str                  # the Excel workbook file name
    sheets: Tuple[str, ...]        # expected worksheet tabs, in order
    csv_columns: Dict[str, List[str]] = field(default_factory=dict)  # csv -> cols

    def expected_csvs(self) -> List[str]:
        """Return the CSV file names this property is expected to write, sorted.

        Phase 8 / S3: every CSV now lives in the property's ``csv/`` subfolder,
        so this is the exact set the subfolder must hold, with nothing extra and
        nothing missing. The workbook is checked separately at the property root.
        """
        return sorted(self.csv_columns.keys())


# The three sample properties, one of each kind, covering the full matrix of
# output shapes the engine can produce today.
PROPERTY_SHAPES: List[PropertyShape] = [
    PropertyShape(
        slug="property-a",
        kind_label="investment (mortgage + rental tax)",
        inputs=DATA_SAMPLE / "property_a" / "inputs.sample.yaml",
        actuals=DATA_SAMPLE / "property_a" / "actuals.sample.csv",
        workbook="property-a_model.xlsx",
        # Phase 8 / S2: workbook renamed to <slug>_model.xlsx; Summary now leads,
        # a dedicated Valuation sheet follows, then the detail and tax sheets.
        sheets=("Summary", "Valuation", "Monthly", "Reconcile", "EventsDaily", "TaxYear", "TenancyLog", "TaxAudit"),
        csv_columns={
            "schedule_monthly.csv": MONTHLY_COLUMNS,
            "reconcile.csv": RECONCILE_COLUMNS,
            "events_daily.csv": EVENTS_DAILY_COLUMNS,
            "tax_year.csv": TAX_YEAR_COLUMNS,
            "tax_audit.csv": TAX_AUDIT_COLUMNS,
        },
    ),
    PropertyShape(
        slug="property-b",
        kind_label="primary residence (mortgage, tax off)",
        inputs=DATA_SAMPLE / "property_b" / "inputs.sample.yaml",
        actuals=DATA_SAMPLE / "property_b" / "actuals.sample.csv",
        workbook="property-b_model.xlsx",
        # Phase 8 / S2: workbook renamed to <slug>_model.xlsx; Summary leads and
        # a Valuation sheet follows. Tax is off, so the three tax sheets stay absent.
        sheets=("Summary", "Valuation", "Monthly", "Reconcile", "EventsDaily"),
        csv_columns={
            "schedule_monthly.csv": MONTHLY_COLUMNS,
            "reconcile.csv": RECONCILE_COLUMNS,
            "events_daily.csv": EVENTS_DAILY_COLUMNS,
        },
    ),
    PropertyShape(
        slug="property-c",
        kind_label="owned outright (valuation-only, no loan)",
        inputs=DATA_SAMPLE / "property_c" / "inputs.sample.yaml",
        # No loan means no bank actuals; omitting --actuals is what routes the
        # engine to its valuation-only path.
        actuals=None,
        workbook="property-c_model.xlsx",
        # Phase 8 / S2: workbook renamed to <slug>_model.xlsx; Summary now leads.
        sheets=("Summary", "Valuation"),
        csv_columns={
            "valuation_schedule.csv": VALUATION_SCHEDULE_COLUMNS,
        },
    ),
]

# Index the shapes by slug so the per-CSV test can look its property back up.
SHAPES_BY_SLUG: Dict[str, PropertyShape] = {s.slug: s for s in PROPERTY_SHAPES}


def _utf8_child_env() -> Dict[str, str]:
    """Return the current environment forced to UTF-8 stdio for a child run.

    Under pytest a subprocess writes to a pipe, which on Windows defaults to the
    legacy code page (cp1252) and cannot encode characters the CLIs print (such
    as the right-arrow in status lines), which would crash the run with a
    UnicodeEncodeError. Forcing UTF-8 keeps the child portable, exactly as the
    golden-master and output-knobs tests do.
    """
    env = dict(os.environ)
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    return env


def _run_engine_for(shape: PropertyShape, out_dir: Path) -> None:
    """Run ``python -m src.engine`` once for a single sample property.

    Finance note: this shells out to exactly the command a person would type by
    hand, so the per-property files this test inspects are the same files a real
    run produces. A valuation-only property has no bank loan to reconcile, so
    ``--actuals`` is left off and the engine takes its no-mortgage path.
    """
    # --inputs and --out are always present; --out is an explicit absolute temp
    # path so the run never depends on a developer's paths.local.yaml.
    cmd = [
        sys.executable, "-m", "src.engine",
        "--inputs", str(shape.inputs),
        "--out", str(out_dir),
    ]
    # --actuals only for a mortgage property; omitting it is the valuation-only
    # trigger the engine branches on.
    if shape.actuals is not None:
        cmd += ["--actuals", str(shape.actuals)]
    subprocess.run(
        cmd,
        cwd=str(REPO_ROOT),  # repo root must be importable under -m
        env=_utf8_child_env(),
        check=True,  # raise immediately if the engine exits non-zero
    )


def _run_portfolio(out_dir: Path) -> None:
    """Regenerate the portfolio rollup into ``out_dir`` via the real CLIs.

    Mirrors the golden-master setup: baseline first, then portfolio, both
    pointed at the bundled sample portfolio and the temp out root, with the data
    and out locations forced through the environment so the run is independent
    of any local config. Only Property A is enabled in the sample portfolio, so
    the rollup pins a single mortgage row.
    """
    env = _utf8_child_env()
    # Force the data and out roots so the run never reads a developer's config.
    env["MORTGAGE_DATA_DIR"] = str(DATA_SAMPLE)
    env["MORTGAGE_OUT_DIR"] = str(out_dir)
    for module in ("tools.baseline", "tools.portfolio"):
        subprocess.run(
            [
                sys.executable, "-m", module,
                "--portfolio", str(SAMPLE_PORTFOLIO),
                "--out", str(out_dir),
            ],
            cwd=str(REPO_ROOT),  # repo root must be importable under -m
            env=env,
            check=True,
        )


@pytest.fixture(scope="session")
def property_outputs(tmp_path_factory: pytest.TempPathFactory) -> Dict[str, Path]:
    """Run each sample property once per session; return slug -> output dir.

    Each property runs into its own throwaway folder so the file-set assertion
    can be exact (nothing from another property leaks in).
    """
    dirs: Dict[str, Path] = {}
    for shape in PROPERTY_SHAPES:
        # A dedicated temp folder per property keeps the produced file set clean.
        out_dir = tmp_path_factory.mktemp(shape.slug)
        _run_engine_for(shape, out_dir)
        dirs[shape.slug] = out_dir
    return dirs


@pytest.fixture(scope="session")
def portfolio_output(tmp_path_factory: pytest.TempPathFactory) -> Path:
    """Run the portfolio rollup once per session and return its output root."""
    out_dir = tmp_path_factory.mktemp("portfolio_rollup")
    _run_portfolio(out_dir)
    return out_dir


def _read_csv_columns(csv_path: Path) -> List[str]:
    """Return the ordered column headings of a CSV without loading its rows.

    Reading zero rows is enough to inspect the header and keeps the check fast
    and immune to value formatting.
    """
    return pd.read_csv(csv_path, nrows=0).columns.tolist()


def _read_sheet_names(workbook_path: Path) -> Tuple[str, ...]:
    """Return the worksheet tab names of an Excel workbook, in order.

    Read-only mode avoids loading cell values, since only the tab inventory and
    its order are part of the locked layout.
    """
    wb = load_workbook(workbook_path, read_only=True)
    try:
        return tuple(wb.sheetnames)
    finally:
        # Always release the file handle, even if reading the names raised.
        wb.close()


@pytest.mark.parametrize("shape", PROPERTY_SHAPES, ids=[s.slug for s in PROPERTY_SHAPES])
def test_property_file_set(property_outputs: Dict[str, Path], shape: PropertyShape) -> None:
    """Each property writes exactly its workbook at the root and its CSVs in csv/.

    A surprise extra file or a missing one is a layout change, so both the
    property-root set (the workbook only) and the csv/ set (exactly the CSVs)
    must equal their locked sets exactly.

    Phase 8 / S3: the CSVs moved down into csv/, so the root now holds only the
    workbook and the csv/ subfolder holds the property's CSVs.
    """
    out_dir = property_outputs[shape.slug]
    # The property root must now hold exactly the workbook; every CSV moved into
    # csv/, so a CSV still sitting at the root is a regression we want to catch.
    produced_root = sorted(p.name for p in out_dir.iterdir() if p.is_file())
    expected_root = [shape.workbook]
    assert produced_root == expected_root, (
        f"{shape.slug} ({shape.kind_label}): the property-root file set changed.\n"
        f"  Expected: {expected_root}\n"
        f"  Produced: {produced_root}\n"
        f"  Missing now:    {[f for f in expected_root if f not in produced_root] or 'none'}\n"
        f"  New/unexpected: {[f for f in produced_root if f not in expected_root] or 'none'}"
    )
    # The csv/ subfolder must hold exactly this property's CSVs, nothing else.
    csv_dir = out_dir / CSV_SUBDIR
    expected_csvs = shape.expected_csvs()
    produced_csvs = (
        sorted(p.name for p in csv_dir.iterdir() if p.is_file())
        if csv_dir.is_dir()
        else []
    )
    assert produced_csvs == expected_csvs, (
        f"{shape.slug} ({shape.kind_label}): the csv/ file set changed.\n"
        f"  Expected: {expected_csvs}\n"
        f"  Produced: {produced_csvs}\n"
        f"  Missing now:    {[f for f in expected_csvs if f not in produced_csvs] or 'none'}\n"
        f"  New/unexpected: {[f for f in produced_csvs if f not in expected_csvs] or 'none'}"
    )


@pytest.mark.parametrize("shape", PROPERTY_SHAPES, ids=[s.slug for s in PROPERTY_SHAPES])
def test_property_workbook_sheets(property_outputs: Dict[str, Path], shape: PropertyShape) -> None:
    """Each property workbook has exactly the expected sheet tabs, in order."""
    workbook_path = property_outputs[shape.slug] / shape.workbook
    assert workbook_path.exists(), (
        f"{shape.slug} ({shape.kind_label}): expected workbook {shape.workbook} "
        f"was not produced at {workbook_path}."
    )
    produced = _read_sheet_names(workbook_path)
    assert produced == shape.sheets, (
        f"{shape.slug} ({shape.kind_label}): {shape.workbook} sheet tabs changed.\n"
        f"  Expected: {list(shape.sheets)}\n"
        f"  Produced: {list(produced)}"
    )


# Flatten the per-property CSV expectations into one case per CSV so a single
# changed file points straight at the property and file that moved.
_CSV_CASES: List[Tuple[str, str, List[str]]] = [
    (shape.slug, csv_name, columns)
    for shape in PROPERTY_SHAPES
    for csv_name, columns in shape.csv_columns.items()
]


@pytest.mark.parametrize(
    "slug,csv_name,expected_columns",
    _CSV_CASES,
    ids=[f"{slug}:{csv_name}" for slug, csv_name, _ in _CSV_CASES],
)
def test_property_csv_columns(
    property_outputs: Dict[str, Path],
    slug: str,
    csv_name: str,
    expected_columns: List[str],
) -> None:
    """Each per-property CSV has exactly the locked columns, in order."""
    # Phase 8 / S3: per-property CSVs now live in the property's csv/ subfolder.
    csv_path = property_outputs[slug] / CSV_SUBDIR / csv_name
    assert csv_path.exists(), (
        f"{slug}: expected CSV {csv_name} was not produced at {csv_path}."
    )
    produced = _read_csv_columns(csv_path)
    kind_label = SHAPES_BY_SLUG[slug].kind_label
    assert produced == expected_columns, (
        f"{slug} ({kind_label}): {csv_name} columns changed.\n"
        f"  Expected: {expected_columns}\n"
        f"  Produced: {produced}\n"
        f"  Missing now:    {[c for c in expected_columns if c not in produced] or 'none'}\n"
        f"  New/unexpected: {[c for c in produced if c not in expected_columns] or 'none'}"
    )


def test_portfolio_summary_csv_columns(portfolio_output: Path) -> None:
    """The portfolio rollup CSV has exactly the locked columns, in order."""
    # Phase 8 / S3: the rollup CSV now lives under a top-level csv/ subfolder;
    # portfolio_summary.xlsx still sits at the out root.
    csv_path = portfolio_output / CSV_SUBDIR / "portfolio_summary.csv"
    assert csv_path.exists(), (
        f"portfolio_summary.csv was not produced at {csv_path}."
    )
    produced = _read_csv_columns(csv_path)
    assert produced == PORTFOLIO_SUMMARY_COLUMNS, (
        "portfolio_summary.csv columns changed.\n"
        f"  Expected: {PORTFOLIO_SUMMARY_COLUMNS}\n"
        f"  Produced: {produced}\n"
        f"  Missing now:    {[c for c in PORTFOLIO_SUMMARY_COLUMNS if c not in produced] or 'none'}\n"
        f"  New/unexpected: {[c for c in produced if c not in PORTFOLIO_SUMMARY_COLUMNS] or 'none'}"
    )


def test_portfolio_summary_workbook_sheet(portfolio_output: Path) -> None:
    """The portfolio rollup workbook has exactly one sheet, named Portfolio."""
    workbook_path = portfolio_output / "portfolio_summary.xlsx"
    assert workbook_path.exists(), (
        f"portfolio_summary.xlsx was not produced at {workbook_path}."
    )
    produced = _read_sheet_names(workbook_path)
    assert produced == ("Portfolio",), (
        "portfolio_summary.xlsx sheet tabs changed.\n"
        f"  Expected: ['Portfolio']\n"
        f"  Produced: {list(produced)}"
    )