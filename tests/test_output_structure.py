# tests/test_output_structure.py
"""Phase 8 / S2: structural contract for the per-property workbooks.

Finance-readable summary
------------------------
This test guards the shape of each property's Excel workbook after the S2
restructure: every property must produce a workbook named after the property
(``<slug>_model.xlsx``) with the right tabs, in the right order, for that kind.
An investment let and a home both get the headline Summary first, then a
Valuation view, then the detail sheets; an owned-outright property gets just
Summary and Valuation. If a later change drops, renames, or reorders a tab, this
test fails and names what moved, so the file a person opens never silently
changes shape.

Technical summary
-----------------
Runs ``python -m src.engine`` once per bundled sample property into its own temp
folder (A and B pass ``--actuals`` and take the mortgage path; C omits it and
the engine branches to the valuation-only path), then asserts the produced
workbook file name and the openpyxl sheet-name order per property kind. Shape
only: it asserts no numbers, so it never overlaps the golden-master numeric
contract. This is the permanent per-kind contract; the v1.8.0 characterization
test remains the transitional full-snapshot net until the S5 re-baseline.
"""
from __future__ import annotations

import os
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple

import pytest
from openpyxl import load_workbook


def _find_repo_root(start: Path) -> Path:
    """Return the repo root by walking up from this file.

    The repo root is the first ancestor holding both ``tools`` and
    ``data_sample`` side by side, mirroring the other suites' resolver so they
    all agree on the root no matter where the file sits.
    """
    # Check this file's own directory first, then each parent in turn.
    for candidate in (start, *start.parents):
        # Only the repo root holds both of these together.
        if (candidate / "tools").is_dir() and (candidate / "data_sample").is_dir():
            return candidate
    # Fail loudly rather than silently resolving to the wrong directory.
    raise RuntimeError(
        "Could not locate the mortflow repo root: no ancestor of "
        f"{start} contains both tools/ and data_sample/."
    )


# Resolve the repo root from this file's location, robust to where it sits.
REPO_ROOT = _find_repo_root(Path(__file__).resolve().parent)
# The bundled de-identified sample data the whole suite runs against.
DATA_SAMPLE = REPO_ROOT / "data_sample"


@dataclass(frozen=True)
class KindShape:
    """The expected workbook name and sheet order for one property kind.

    Finance note: groups, for one property kind, the workbook file name we
    expect and the worksheet tabs that workbook must carry, in order.
    """

    slug: str                  # output sub-folder and workbook slug
    kind_label: str            # plain-language kind, for readable failures
    inputs: Path               # the property's sample inputs YAML
    actuals: Optional[Path]    # bank actuals CSV, or None for valuation-only
    workbook: str              # expected <slug>_model.xlsx name
    sheets: Tuple[str, ...]    # expected worksheet tabs, in order


# One property per kind, covering the whole output matrix the engine can produce.
KIND_SHAPES = [
    KindShape(
        slug="property-a",
        kind_label="investment (mortgage + rental tax)",
        inputs=DATA_SAMPLE / "property_a" / "inputs.sample.yaml",
        actuals=DATA_SAMPLE / "property_a" / "actuals.sample.csv",
        workbook="property-a_model.xlsx",
        # Summary first, then the Valuation view, then the detail sheets, then
        # the tax sheets (rental tax is on for an investment let).
        sheets=("Summary", "Valuation", "Monthly", "Reconcile", "EventsDaily", "TaxYear", "TenancyLog", "TaxAudit"),
    ),
    KindShape(
        slug="property-b",
        kind_label="primary residence (mortgage, tax off)",
        inputs=DATA_SAMPLE / "property_b" / "inputs.sample.yaml",
        actuals=DATA_SAMPLE / "property_b" / "actuals.sample.csv",
        workbook="property-b_model.xlsx",
        # Same mortgage shape as an investment, minus the three tax sheets.
        sheets=("Summary", "Valuation", "Monthly", "Reconcile", "EventsDaily"),
    ),
    KindShape(
        slug="property-c",
        kind_label="owned outright (valuation-only, no loan)",
        inputs=DATA_SAMPLE / "property_c" / "inputs.sample.yaml",
        # No loan means no bank actuals; omitting --actuals routes the engine to
        # its valuation-only path.
        actuals=None,
        workbook="property-c_model.xlsx",
        # No loan: just the headline and the value-over-time view.
        sheets=("Summary", "Valuation"),
    ),
]


def _utf8_child_env() -> Dict[str, str]:
    """Return the current environment forced to UTF-8 stdio for a child run.

    Under pytest a subprocess writes to a pipe, which on Windows defaults to the
    legacy code page (cp1252) and cannot encode characters the CLI prints (such
    as the right-arrow in status lines), which would crash the run with a
    UnicodeEncodeError. Forcing UTF-8 keeps the child portable, exactly as the
    golden-master and characterization tests do.
    """
    env = dict(os.environ)
    env["PYTHONUTF8"] = "1"
    env["PYTHONIOENCODING"] = "utf-8"
    return env


def _run_engine_for(shape: KindShape, out_dir: Path) -> None:
    """Run ``python -m src.engine`` once for a single sample property.

    Finance note: shells out to exactly the command a person would type by hand,
    so the workbook this test inspects is the same file a real run produces. A
    valuation-only property has no bank loan to reconcile, so ``--actuals`` is
    left off and the engine takes its no-mortgage path.
    """
    cmd = [
        sys.executable, "-m", "src.engine",
        "--inputs", str(shape.inputs),
        "--out", str(out_dir),
    ]
    # --actuals only for a mortgage property; omitting it is the valuation-only
    # trigger the engine branches on.
    if shape.actuals is not None:
        cmd += ["--actuals", str(shape.actuals)]
    subprocess.run(
        cmd,
        cwd=str(REPO_ROOT),  # repo root must be importable under -m
        env=_utf8_child_env(),
        check=True,  # raise immediately if the engine exits non-zero
    )


@pytest.fixture(scope="session")
def kind_outputs(tmp_path_factory: pytest.TempPathFactory) -> Dict[str, Path]:
    """Run each sample property once per session; return slug -> output dir.

    Each property runs into its own throwaway folder so the workbook check sees
    only that property's files.
    """
    dirs: Dict[str, Path] = {}
    for shape in KIND_SHAPES:
        out_dir = tmp_path_factory.mktemp(shape.slug)
        _run_engine_for(shape, out_dir)
        dirs[shape.slug] = out_dir
    return dirs


def _read_sheet_names(workbook_path: Path) -> Tuple[str, ...]:
    """Return the worksheet tab names of a workbook, in order (read-only)."""
    wb = load_workbook(workbook_path, read_only=True)
    try:
        return tuple(wb.sheetnames)
    finally:
        # Always release the file handle, even if reading the names raised.
        wb.close()


@pytest.mark.parametrize("shape", KIND_SHAPES, ids=[s.slug for s in KIND_SHAPES])
def test_workbook_named_for_property(kind_outputs: Dict[str, Path], shape: KindShape) -> None:
    """Each property writes exactly one ``<slug>_model.xlsx`` workbook."""
    out_dir = kind_outputs[shape.slug]
    # Only Excel files matter here; CSV demotion is checked elsewhere.
    workbooks = sorted(p.name for p in out_dir.iterdir() if p.is_file() and p.suffix == ".xlsx")
    assert workbooks == [shape.workbook], (
        f"{shape.slug} ({shape.kind_label}): the workbook name changed.\n"
        f"  Expected: ['{shape.workbook}']\n"
        f"  Produced: {workbooks}"
    )


@pytest.mark.parametrize("shape", KIND_SHAPES, ids=[s.slug for s in KIND_SHAPES])
def test_workbook_sheet_set_and_order(kind_outputs: Dict[str, Path], shape: KindShape) -> None:
    """Each property workbook has exactly the expected tabs, in order.

    Summary leads every workbook and a Valuation view follows; a missing or
    reordered tab is a layout regression and fails here.
    """
    workbook_path = kind_outputs[shape.slug] / shape.workbook
    assert workbook_path.exists(), (
        f"{shape.slug} ({shape.kind_label}): expected workbook {shape.workbook} "
        f"was not produced at {workbook_path}."
    )
    produced = _read_sheet_names(workbook_path)
    assert produced == shape.sheets, (
        f"{shape.slug} ({shape.kind_label}): {shape.workbook} sheet tabs changed.\n"
        f"  Expected: {list(shape.sheets)}\n"
        f"  Produced: {list(produced)}"
    )