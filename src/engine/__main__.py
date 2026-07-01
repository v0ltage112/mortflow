# src/engine/__main__.py
"""Command-line entry point for the mortgage engine package.

Finance-readable summary
------------------------
This is the runnable wrapper a person (or another script) uses to produce the
actual output files. It loads the config and the bank statement, runs the
engine once, builds the headline Summary figures, and writes the Excel workbook
and CSVs (Monthly, Reconcile, Events, and the optional Tax sheets). It does no
mortgage maths of its own; it orchestrates the pieces and saves the results so
the numbers can be reviewed.

Technical summary
-----------------
Defines ``main()`` and the ``python -m src.engine`` guard. Loads inputs/actuals,
calls ``run_engine``, derives summary/portal metrics, and writes XLSX + CSV
artefacts (plus optional tax outputs).

Phase 5 / S1 note: ``main()`` was relocated verbatim from the original
``src/engine.py`` CLI section, raising the optional tax/paths import depth from
``.tax``/``.paths`` to ``..tax``/``..paths`` because this module sits one level
deeper (``src/engine/__main__.py``); ``..tax``/``..paths`` still resolve to
``src.tax``/``src.paths`` and keep ``python -m src.engine`` working.

Phase 5 / S6 note: the dead-import noise is gone and the property-value growth
coercion is routed through ``helpers.growth_to_decimal``.

Phase 6 / S3 note: the CLI branches on the property kind. A no-mortgage property
(``meta.mortgage_enabled`` False) is routed to ``valuation_only.run_valuation_only``;
``--actuals`` is optional at the parser level but still required for a mortgage
property. Mortgage-bearing properties run exactly as before.

Phase 6 / S4 note: the writer path now honours the parsed ``output`` block.
``write_excel`` gates the workbook, ``write_csv`` gates the CSV artefacts,
``include_daily_events`` gates both the EventsDaily sheet and ``events_daily.csv``,
and ``currency`` selects the money mask via ``report.money_number_format``. Every
default is on with euro formatting, so an unchanged config writes byte-identical
outputs and the Gandon golden master stays green. The stdout "Wrote outputs to:"
line is unchanged.

Phase 8 / S2 note: the workbook is renamed and restructured. It is now named
``<slug>_model.xlsx`` (the slug is the same one the output folder uses, via
``helpers.slugify``) instead of the generic ``mortgage_outputs.xlsx``; the
Summary sheet is moved to the front; a dedicated Valuation sheet (month-end
property value and the model and bank LTVs, reused from the Monthly schedule) is
added; and the Phase 7 attribution split totals (contractual, overpayment,
lump, difference) are surfaced on the Summary. These are file-shape and
presentation changes only: the engine maths and every reported number are
unchanged, so the golden master (which checks CSV values) stays green. The CSV
file names are unchanged here; demoting them into a ``csv/`` subfolder is S3.

Phase 8 / S3 note: that demotion is now done. Every CSV the mortgage path writes
(``schedule_monthly``, ``reconcile``, ``events_daily``, and the optional
``tax_year`` / ``tax_audit``) is written into the ``out_cfg.csv_subdir``
sub-folder (default ``csv``) instead of beside the workbook; the
``<slug>_model.xlsx`` workbook stays at the property root. An empty
``csv_subdir`` restores the flat layout. CSV contents and the engine maths are
unchanged, so the golden master (which compares CSV values) stays green; only
the paths moved.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
import yaml
from openpyxl.styles import Font

from .helpers import ensure_date, growth_to_decimal, slugify
from .schema import load_inputs, load_actuals
from .simulate import run_engine
# Phase 6 / S4: money_number_format turns the output.currency knob into the Excel
# money mask; it defaults to the original euro mask for EUR.
from .report import _add_table, _format_sheet, compute_portal_style_metrics, money_number_format
# Phase 6 / S3: the no-mortgage valuation-only path. Imported here (not in
# __init__) to keep the package facade minimal; no circular import because
# valuation_only depends only on helpers/schema/valuation/monthly/report.
from .valuation_only import run_valuation_only

# Tax and path-resolver modules live one level up in ``src``.
from ..tax import load_tenancies, compute_tax_year_table
from ..paths import resolve_out_dir, resolve_relative


def main():
    """Entry point used by ``python -m src.engine`` and ``src/engine.py``.

    Finance note: this is the button that produces the deliverables. It reads
    the configuration and bank statement, runs the model, assembles the Summary
    headline figures (as-of balance, YTD interest/principal, next payment, LTV),
    and writes the Excel workbook and CSVs that the business actually reviews.

    The CLI is intentionally lightweight: load configuration, run the engine,
    and write CSV/XLSX artefacts.  Any consumer that needs to re-use the logic
    should import :func:`run_engine` directly instead of shelling out.
    """
    # Plain-English progress line for troubleshooting (stderr only; the stdout
    # "Wrote outputs to:" line below stays byte-identical for the golden master).
    print("[engine.__main__] starting CLI run", file=sys.stderr)

    ap = argparse.ArgumentParser(description="Daily mortgage engine (ACT/365)")
    ap.add_argument("--inputs", type=Path, required=True, help="Path to inputs.yaml")
    # --actuals is optional at the parser level so a no-mortgage (valuation-only)
    # property can omit it. A mortgage-enabled property that omits it still
    # errors below, preserving the original contract for the loan path.
    ap.add_argument("--actuals", type=Path, required=False, default=None, help="Path to actuals.csv (required for a mortgage property)")
    # --out is optional now.  When omitted, the output folder is resolved through
    # the Phase 2 config layer (CLI > MORTGAGE_OUT_DIR > paths.local.yaml > <repo>/out).
    ap.add_argument("--out", type=Path, default=None, help="Output folder (overrides config)")
    args = ap.parse_args()

    # Resolve the output directory through the config layer.  Passing the raw CLI
    # value (or None) keeps an explicit --out as the highest-priority source.
    out_dir = resolve_out_dir(str(args.out) if args.out is not None else None)
    out_dir.mkdir(parents=True, exist_ok=True)

    inputs = load_inputs(args.inputs)

    # Phase 6 / S3: a no-mortgage property has no loan schedule, reconcile, or
    # tax to run. Branch to the valuation-only path, which writes a minimal
    # value-over-time output and returns. Mortgage-bearing properties skip this
    # block entirely and run exactly as before (the Gandon golden master path).
    meta = inputs.meta
    if meta is not None and not meta.mortgage_enabled:
        if not meta.valuation_enabled:
            # Mortgage off and valuation off means there is nothing to compute.
            ap.error("property has neither mortgage nor valuation enabled; nothing to run")
        run_valuation_only(inputs, args.inputs, out_dir)
        # Completion line mirrors the mortgage path for consistent troubleshooting.
        print("[engine.__main__] CLI run complete", file=sys.stderr)
        return

    # A mortgage-enabled property must have a bank-actuals CSV to reconcile
    # against; --actuals is optional at the parser level only so the valuation-
    # only path above can omit it.
    if args.actuals is None:
        ap.error("--actuals is required for a mortgage-enabled property")

    actuals = load_actuals(args.actuals)

    # Core engine run ---------------------------------------------------------
    monthly, reconcile, events = run_engine(inputs, actuals)

    # ---- Derive \"as-of\" and quick summary stats for Summary sheet
    rec_non_na = (
        reconcile.dropna(subset=["model_balance"]) if "model_balance" in reconcile.columns else reconcile.copy()
    )
    asof_date = rec_non_na["bank_date"].max() if not rec_non_na.empty else None

    bank_bal = None
    model_bal_same = None
    diff_bal = None
    if asof_date is not None and "bank_running_balance" in rec_non_na.columns:
        bank_row = rec_non_na.loc[rec_non_na["bank_date"] == asof_date]
        if not bank_row.empty:
            bank_bal = float(bank_row["bank_running_balance"].iloc[-1])
            model_bal_same = float(bank_row["model_balance"].iloc[-1]) if "model_balance" in bank_row.columns else None
            if model_bal_same is not None:
                diff_bal = model_bal_same - bank_bal

    # Optional: prefer a newer portal snapshot if present
    raw_cfg = yaml.safe_load(Path(args.inputs).read_text())
    snaps = (raw_cfg.get("reconcile") or {}).get("snapshots") or []
    if snaps:
        latest = max(snaps, key=lambda s: ensure_date(s["date"]))
        snap_dt = ensure_date(latest["date"])
        if asof_date is None or snap_dt > asof_date:
            asof_date = snap_dt
            bank_bal = float(latest["balance"])
            portal_metrics = compute_portal_style_metrics(asof_date, inputs, events, monthly)
            model_bal_same = portal_metrics.get("principal_excl_unposted")
            diff_bal = (model_bal_same - bank_bal) if (model_bal_same is not None) else None

    # YTD (calendar-year posting)
    ytd_interest = None
    ytd_principal = None
    if asof_date is not None and "posting_date" in monthly.columns:
        md = monthly.copy()
        md["posting_date"] = pd.to_datetime(md["posting_date"])
        cond = (md["posting_date"].dt.year == pd.Timestamp(asof_date).year) & (
            md["posting_date"] <= pd.Timestamp(asof_date)
        )
        ytd_interest = float(md.loc[cond, "interest_used"].sum())
        ytd_principal = float(md.loc[cond, "principal_paid"].sum())

    # Next scheduled payment (from event log)
    next_pay_date = None
    next_pay_amt = None
    if asof_date is not None:
        future_pays = events[(events["kind"] == "Payment") & (events["date"] > asof_date)]
        if not future_pays.empty:
            row = future_pays.sort_values("date").iloc[0]
            next_pay_date = row["date"]
            next_pay_amt = float(row["amount"])

    # Current annual rate at as-of
    cur_rate = None
    if asof_date is not None and "ym" in monthly.columns and "annual_rate" in monthly.columns:
        ym_key = int(pd.Timestamp(asof_date).year * 100 + pd.Timestamp(asof_date).month)
        hit = monthly.loc[monthly["ym"] == ym_key]
        if not hit.empty:
            cur_rate = float(hit.iloc[0]["annual_rate"])

    # Property value path (coerce 1.00 to 1% if user wrote percent)
    prop_val = None
    ltv = None
    try:
        # Normalise the growth assumption through the shared helper so the CLI
        # summary uses the same whole-percent-vs-decimal rule as the engine.
        growth = growth_to_decimal(inputs.property_growth_pa)
        if asof_date is not None:
            m_since = (pd.Timestamp(asof_date).year - inputs.drawdown_date.year) * 12 + (
                pd.Timestamp(asof_date).month - inputs.drawdown_date.month
            )
            prop_val = float(inputs.property_price * ((1.0 + growth) ** (m_since / 12.0)))
            if bank_bal is not None and prop_val > 0:
                ltv = bank_bal / prop_val
    except Exception:
        pass

    # ---- Tax (optional) ----
    tax_cfg = (raw_cfg.get("tax") or {})
    tax_enabled = bool(tax_cfg.get("enabled", False))

    tax_year_df = None
    ten_log_df = None
    tax_audit_df = None
    tenancies = None
    policy = None

    if tax_enabled:
        # Tax computations are optional and live in ``src.tax``.  They reuse the
        # monthly schedule produced earlier and therefore inherit the same
        # assumptions as the engine.
        pref = resolve_relative(args.inputs, tax_cfg.get("tenancy_file", "tenancy.local.yaml"))
        fb = resolve_relative(args.inputs, "tenancy.sample.yaml")
        tenancies, policy, _ = load_tenancies(pref, fb)
        tax_year_df, ten_log_df = compute_tax_year_table(monthly, raw_cfg, tenancies, policy)

        # monthly audit only when tax is enabled
        if (tax_cfg.get("audit") or {}).get("write_monthly", True):
            # Imported lazily because it is only needed when the monthly tax
            # audit is switched on; ..tax matches the top-of-file imports.
            from ..tax import compute_tax_monthly_audit
            tax_audit_df = compute_tax_monthly_audit(monthly, raw_cfg, tenancies, policy)

    # ---- Write outputs ----
    # Phase 6 / S4: the output artefacts are now gated by the parsed `output`
    # block. Every default is on with euro formatting, so an unchanged config
    # writes exactly the same files it wrote before these knobs were honoured
    # and the Gandon golden master stays green.
    out_cfg = inputs.output
    # Currency selects the money symbol used on every sheet; EUR reproduces the
    # original euro mask character for character, so a default run moves nothing.
    money_fmt = money_number_format(out_cfg.currency)

    # Phase 8 / S2: name the workbook from the property slug, the same slug the
    # output folder uses (via helpers.slugify), so each property produces a
    # clearly named <slug>_model.xlsx (for example property-a_model.xlsx) rather
    # than a generic mortgage_outputs.xlsx. Fall back to the output folder name,
    # then a literal "property", if a config somehow carries no meta name.
    meta_name = inputs.meta.name if (inputs.meta and inputs.meta.name) else None
    slug = slugify(meta_name) if meta_name else (slugify(out_dir.name) or "property")
    workbook_name = f"{slug}_model.xlsx"

    # XLSX (only when the workbook is switched on) ---------------------------
    if out_cfg.write_excel:
        with pd.ExcelWriter(out_dir / workbook_name, engine="openpyxl") as xl:
            monthly.to_excel(xl, sheet_name="Monthly", index=False)
            reconcile.to_excel(xl, sheet_name="Reconcile", index=False)
            # The daily EventsDaily sheet is optional under include_daily_events.
            if out_cfg.include_daily_events:
                events.to_excel(xl, sheet_name="EventsDaily", index=False)

            wb = xl.book
            ws_m = xl.sheets["Monthly"]
            ws_r = xl.sheets["Reconcile"]

            # Tables + formats
            _add_table(ws_m, "Monthly")
            _add_table(ws_r, "Reconcile")

            _format_sheet(
                ws_m,
                money_cols=[
                    # Phase 7 / S4: final attribution vocabulary. payment_amount
                    # and extra_amount are retired; lump_amount is now lump and
                    # payment_unattributed is now difference.
                    "contractual", "overpayment", "lump", "total_paid", "difference",
                    "interest_used", "principal_paid",
                    "model_eom_balance", "bank_eom_running_balance", "eom_diff_model_minus_bank",
                    "property_value"
                ],
                pct_cols=["annual_rate", "ltv_model_eom", "ltv_bank_eom"],
                date_cols=["month_start", "payment_date", "posting_date"],
                money_format=money_fmt,
            )

            _format_sheet(
                ws_r,
                money_cols=["amount", "bank_running_balance", "model_amount", "model_balance", "diff_model_minus_bank"],
                date_cols=["bank_date"],
                money_format=money_fmt,
            )

            # EventsDaily is formatted only when it was written.
            if out_cfg.include_daily_events:
                ws_e = xl.sheets["EventsDaily"]
                _add_table(ws_e, "EventsDaily")
                _format_sheet(
                    ws_e,
                    money_cols=["amount", "balance", "property_value"],
                    pct_cols=["ltv_after_event"],
                    date_cols=["date"],
                    money_format=money_fmt,
                )

            # ---------------- Tax sheets (optional) ----------------
            if tax_enabled and tax_year_df is not None:
                tax_year_df.to_excel(xl, sheet_name="TaxYear", index=False)
                ws_t = xl.sheets["TaxYear"]
                _add_table(ws_t, "TaxYear")
                _format_sheet(
                    ws_t,
                    money_cols=["interest_posted", "allowable_interest_s97", "principal_paid"],
                    pct_cols=["avg_occupancy_ratio", "deductible_pct"],
                    date_cols=[],
                    money_format=money_fmt,
                )

                ten_log_df.to_excel(xl, sheet_name="TenancyLog", index=False)
                ws_log = xl.sheets["TenancyLog"]
                _add_table(ws_log, "TenancyLog")
                _format_sheet(
                    ws_log,
                    money_cols=["rent_amount", "security_deposit"],
                    pct_cols=[],
                    date_cols=["start", "end", "rtb_registration_date"],
                    money_format=money_fmt,
                )
            if tax_enabled and tax_audit_df is not None:
                tax_audit_df.to_excel(xl, sheet_name="TaxAudit", index=False)
                ws_a = xl.sheets["TaxAudit"]
                _add_table(ws_a, "TaxAudit")
                _format_sheet(
                    ws_a,
                    money_cols=["interest_used", "principal_paid", "allowable_interest_s97"],
                    pct_cols=["occupancy_ratio", "deductible_pct"],
                    date_cols=["month_start", "posting_date"],
                    money_format=money_fmt,
                )

            # ---------------- Valuation (mortgaged property) ----------------
            # Phase 8 / S2: a dedicated value and loan-to-value view, so the
            # property-value path is readable on its own tab instead of being
            # buried among the wide Monthly schedule columns. It reuses figures
            # already in the monthly schedule (the month-end property value and
            # the model and bank LTVs), so nothing is recomputed and the numbers
            # match the Monthly sheet exactly.
            valuation_view_cols = ["month_start", "ym", "property_value", "ltv_model_eom", "ltv_bank_eom"]
            # Guard each column so a future schedule that renames or drops one
            # degrades to a thinner sheet rather than crashing the writer.
            valuation_view = monthly[[c for c in valuation_view_cols if c in monthly.columns]].copy()
            valuation_view.to_excel(xl, sheet_name="Valuation", index=False)
            ws_v = xl.sheets["Valuation"]
            _add_table(ws_v, "Valuation")
            _format_sheet(
                ws_v,
                money_cols=["property_value"],
                pct_cols=["ltv_model_eom", "ltv_bank_eom"],
                date_cols=["month_start"],
                money_format=money_fmt,
            )

            # ---------------- Summary (values only) ----------------
            ws_s = wb.create_sheet("Summary")
            ws_s.append(["Metric", "Value"])

            portal = (
                compute_portal_style_metrics(asof_date, inputs, events, monthly)
                if asof_date
                else {"principal_excl_unposted": None, "ytd_interest_portal": None}
            )

            # Phase 8 / S2: Phase 7 attribution split totals for the Summary, so
            # the headline sheet shows how the total paid divides into the
            # contractual amount, the agreed overpayment, any lump, and the
            # unattributed difference. Summed straight from the monthly schedule;
            # these are presentation totals, not new engine figures.
            def _column_total(df, col):
                """Return a monthly column's sum as a float, or None if absent."""
                return float(df[col].sum()) if col in df.columns else None

            total_contractual = _column_total(monthly, "contractual")
            total_overpayment = _column_total(monthly, "overpayment")
            total_lump = _column_total(monthly, "lump")
            total_difference = _column_total(monthly, "difference")

            rows = [
                ("As of date (latest bank actual)", asof_date),
                ("Bank running balance (as-of)", bank_bal),
                ("Model balance same date", model_bal_same),
                ("Difference (model - bank)", diff_bal),
                ("Current annual rate", cur_rate),
                ("Next payment date", next_pay_date),
                ("Next payment amount", next_pay_amt),
                ("YTD interest (posted only)", ytd_interest),
                ("YTD principal (posted only)", ytd_principal),
                # Phase 8 / S2: Phase 7 attribution split totals over all months.
                ("Total contractual paid", total_contractual),
                ("Total overpayment", total_overpayment),
                ("Total lump", total_lump),
                ("Total difference (unattributed)", total_difference),
                ("Portal-style principal (excl. unposted interest)", portal["principal_excl_unposted"]),
                ("Portal YTD interest (posted + accrual to yesterday)", portal["ytd_interest_portal"]),
                ("Property value (as-of)", prop_val),
                ("LTV (as-of)", ltv),
            ]
            for k, v in rows:
                ws_s.append([k, v])

            ws_s.freeze_panes = "A2"
            ws_s["A1"].font = Font(bold=True)
            ws_s["B1"].font = Font(bold=True)
            money_keys = {
                "Bank running balance (as-of)",
                "Model balance same date",
                "Difference (model - bank)",
                "Next payment amount",
                "Property value (as-of)",
                "Portal-style principal (excl. unposted interest)",
                "Portal YTD interest (posted + accrual to yesterday)",
                # Phase 8 / S2: the attribution split totals are money too.
                "Total contractual paid",
                "Total overpayment",
                "Total lump",
                "Total difference (unattributed)",
            }
            pct_keys = {"LTV (as-of)", "Current annual rate"}
            date_keys = {"As of date (latest bank actual)", "Next payment date"}
            for r in range(2, ws_s.max_row + 1):
                k = ws_s.cell(row=r, column=1).value
                v = ws_s.cell(row=r, column=2)
                if k in money_keys:
                    # Currency-aware mask (euro by default) keeps a default run identical.
                    v.number_format = money_fmt
                if k in pct_keys:
                    v.number_format = "0.00%"
                if k in date_keys:
                    v.number_format = "yyyy-mm-dd"
            ws_s.column_dimensions["A"].width = 48
            ws_s.column_dimensions["B"].width = 28

            # Phase 8 / S2: present Summary first, then the Valuation view, then
            # the detail sheets in their existing creation order. Reordering tabs
            # is purely cosmetic; no value or formula is touched.
            lead = ["Summary", "Valuation"]
            wb._sheets = [wb[t] for t in lead if t in wb.sheetnames] + [
                ws for ws in wb._sheets if ws.title not in lead
            ]

    # CSVs (only when CSV output is switched on) -----------------------------
    # Phase 8 / S3: every CSV is demoted into the out_cfg.csv_subdir sub-folder
    # (default "csv") so the property folder shows the headline workbook plus a
    # tidy csv/ folder. An empty csv_subdir restores the old flat layout. The
    # workbook written above is unaffected; it stays at the property root.
    if out_cfg.write_csv:
        csv_dir = (out_dir / out_cfg.csv_subdir) if out_cfg.csv_subdir else out_dir
        csv_dir.mkdir(parents=True, exist_ok=True)
        monthly.to_csv(csv_dir / "schedule_monthly.csv", index=False)
        reconcile.to_csv(csv_dir / "reconcile.csv", index=False)
        # The daily events CSV shares the include_daily_events switch with the
        # workbook sheet so the two artefacts stay consistent.
        if out_cfg.include_daily_events:
            events.to_csv(csv_dir / "events_daily.csv", index=False)
        if tax_enabled and tax_year_df is not None:
            tax_year_df.to_csv(csv_dir / "tax_year.csv", index=False)
        if tax_enabled and tax_audit_df is not None:
            tax_audit_df.to_csv(csv_dir / "tax_audit.csv", index=False)

    print("Wrote outputs to:", out_dir.resolve())
    # Plain-English completion line for troubleshooting (stderr only).
    print("[engine.__main__] CLI run complete", file=sys.stderr)


if __name__ == "__main__":
    main()