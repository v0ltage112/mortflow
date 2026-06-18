# src/engine/valuation_only.py
"""Valuation-only run path for properties with no mortgage.

Finance-readable summary
------------------------
Some properties in the portfolio carry no mortgage at all (a fully owned
apartment, for example). For those there is no loan schedule, no bank
reconciliation, and no rental-tax computation to run. The only thing the model
tracks is what the property is worth over time. This module is that path: it
takes the property's base value, growth assumption, and a base date, and walks
the value forward month by month to the modelling horizon, then writes a small
value-over-time output (a CSV and a formatted workbook). There is no
loan-to-value figure here because there is no loan.

Technical summary
-----------------
Pure, side-effect-free build plus a thin writer. ``_read_valuation_anchor``
collects the minimal value inputs from a dedicated top-level ``valuation:``
block (with a fall back to any loan-derived fields already on ``Inputs``).
``build_valuation_schedule`` walks ``monthly.month_span`` and values each month
through the shared ``valuation.property_value_on`` so the growth/revaluation
maths has exactly one definition across the package. ``write_valuation_outputs``
saves ``valuation_schedule.csv`` and a formatted ``valuation_outputs.xlsx`` via
the shared ``report`` helpers. ``run_valuation_only`` orchestrates the three.

Phase 6 / S3 note: new module. The CLI (``src/engine/__main__.py``) branches to
``run_valuation_only`` when ``meta.mortgage_enabled`` is False, so a no-mortgage
property runs end to end instead of crashing in the loan engine. Mortgage-
bearing properties never enter this path and are byte-for-byte unchanged.

Phase 6 / S4 note: the writer now honours the parsed ``output`` block. ``write_csv``
gates ``valuation_schedule.csv``; ``write_excel`` gates ``valuation_outputs.xlsx``;
``currency`` selects the money mask via ``report.money_number_format``.
``include_daily_events`` does not apply because there is no daily event log on
this path. Every default is on with euro formatting, so a default valuation-only
run is identical to S3.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import List

import pandas as pd
import yaml
from openpyxl.styles import Font

from .helpers import ensure_date, growth_to_decimal
from .schema import Inputs, ValuationBlock
from .valuation import property_value_on
from .monthly import month_span
# Phase 6 / S4: money_number_format turns the output.currency knob into the Excel
# money mask; it defaults to the original euro mask for EUR.
from .report import _add_table, _format_sheet, money_number_format


@dataclass
class _ValuationAnchor:
    """The minimal value inputs ``property_value_on`` needs, for a no-loan property.

    Finance note: an owned-outright property has no ``loan`` block, so the price,
    growth rate, and the date the value is anchored from come from a dedicated
    ``valuation`` block instead. This small object carries exactly those fields
    under the same attribute names the shared valuation helper reads, so the
    value-over-time maths is identical to a mortgaged property's LTV path.
    """

    # Attribute names deliberately mirror ``Inputs`` so ``property_value_on`` can
    # consume this object unchanged (it only reads these four attributes).
    property_price: float            # base value the growth path starts from
    property_growth_pa: float        # annual growth (decimal or whole percent)
    drawdown_date: date              # reused as the base-valuation date
    valuation_blocks: List[ValuationBlock] = field(default_factory=list)


def _read_valuation_anchor(inputs: Inputs, raw_cfg: dict) -> _ValuationAnchor:
    """Collect the base value, growth, base date, and any revaluation blocks.

    Finance note: reads the dedicated ``valuation`` block an owned-outright
    config carries (base value, growth, the date the value is anchored from, and
    optional later revaluations). If that block is absent it falls back to any
    loan-derived value already on ``Inputs`` so a mortgaged config could in
    principle reuse the same path. It fails loudly when there is no positive base
    value or no base date, because then there is nothing meaningful to grow.
    """
    val = (raw_cfg.get("valuation") or {})

    # Base value: dedicated block first ('base_value' or 'value'), then any
    # loan-derived price the schema already parsed onto Inputs.
    base_value = val.get("base_value", val.get("value"))
    if base_value is None:
        base_value = inputs.property_price
    base_value = float(base_value or 0.0)

    # Base date: dedicated block first ('base_date' or 'start'), then the loan
    # drawdown date if a loan block happened to be present.
    base_date_raw = val.get("base_date", val.get("start"))
    if base_date_raw is not None:
        base_date = ensure_date(base_date_raw)
    elif inputs.drawdown_date is not None:
        base_date = inputs.drawdown_date
    else:
        base_date = None

    # Growth: dedicated block first, then the loan-derived growth on Inputs.
    # property_value_on normalises this through growth_to_decimal itself, so the
    # raw user value (0.03 or 3) is passed straight through here.
    growth_raw = val.get("growth_pa", inputs.property_growth_pa)

    # Optional revaluation blocks, normalised exactly like schema.load_inputs so
    # a value/growth regime change behaves the same whether the property has a
    # loan or not.
    blocks_raw = (val.get("blocks") or val.get("valuation_blocks") or [])
    blocks: List[ValuationBlock] = []
    for vb in blocks_raw:
        blocks.append(
            ValuationBlock(
                start=ensure_date(vb["start"]),
                base_value=float(vb["value"]),
                growth_pa=growth_to_decimal(vb.get("growth_pa", growth_raw)),
            )
        )
    # Inherit any blocks the schema already parsed from a loan block, if present.
    if inputs.valuation_blocks:
        blocks.extend(inputs.valuation_blocks)
    blocks = sorted(blocks, key=lambda b: b.start)

    # Nothing to grow without a positive base value and a base date: fail loud
    # rather than silently emitting a flat or empty series.
    if base_value <= 0.0 or base_date is None:
        raise ValueError(
            "valuation-only run needs a positive base value and a base date; "
            "add a 'valuation' block with base_value and base_date (or a loan "
            "block carrying property_price and drawdown_date)"
        )

    return _ValuationAnchor(
        property_price=base_value,
        property_growth_pa=growth_raw,
        drawdown_date=base_date,
        valuation_blocks=blocks,
    )


def build_valuation_schedule(inputs: Inputs, anchor: _ValuationAnchor) -> pd.DataFrame:
    """Return a month-by-month property value series to the modelling horizon.

    Finance note: this is the whole deliverable for a no-mortgage property: one
    row per calendar month from the base-valuation date to the modelling end
    date, each showing the modelled property value that month. There is no
    balance and no LTV because there is no loan.
    """
    end = inputs.modelling_end_date
    if end is None:
        # Without a horizon there is no end to the value series; the contractual
        # term fallback used for loans does not exist for an owned-outright file.
        raise ValueError(
            "valuation-only run needs a modelling horizon; add modelling.end_date"
        )

    # One first-of-month entry from the base date to the horizon (shared helper).
    months = month_span(anchor.drawdown_date, end)

    rows = []
    for ms in months:
        rows.append(
            dict(
                month_start=ms,
                ym=ms.year * 100 + ms.month,
                # Reuse the shared valuation maths; the anchor duck-types as Inputs.
                property_value=round(property_value_on(anchor, ms), 2),
            )
        )
    return pd.DataFrame(rows)


def write_valuation_outputs(
    out_dir: Path, schedule: pd.DataFrame, inputs: Inputs, anchor: _ValuationAnchor
) -> None:
    """Write the valuation-only CSV and a formatted XLSX workbook.

    Finance note: saves the value-over-time table as a plain CSV (for machines
    and fixtures) and as a readable Excel workbook with a Valuation sheet and a
    one-look Summary sheet (base value, growth, horizon value). Formatting only;
    no figure is changed here.

    Phase 6 / S4: the ``output`` block now gates the two artefacts and selects
    the money mask. Defaults are on with euro formatting, so a default run is
    identical to S3. ``include_daily_events`` is not consulted because this path
    has no daily event log.
    """
    # Phase 6 / S4: honour the output knobs. write_csv and write_excel gate the
    # two artefacts; currency selects the money symbol used in the workbook.
    out_cfg = inputs.output
    money_fmt = money_number_format(out_cfg.currency)

    # CSV first: machine-readable and friendly to a future locked fixture.
    if out_cfg.write_csv:
        schedule.to_csv(out_dir / "valuation_schedule.csv", index=False)

    if out_cfg.write_excel:
        with pd.ExcelWriter(out_dir / "valuation_outputs.xlsx", engine="openpyxl") as xl:
            schedule.to_excel(xl, sheet_name="Valuation", index=False)
            wb = xl.book
            ws_v = xl.sheets["Valuation"]

            # Table + number formats reuse the shared report helpers for one look.
            _add_table(ws_v, "Valuation")
            _format_sheet(
                ws_v,
                money_cols=["property_value"],
                date_cols=["month_start"],
                money_format=money_fmt,
            )

            # ---------------- Summary (values only) ----------------
            ws_s = wb.create_sheet("Summary")
            ws_s.append(["Metric", "Value"])

            first_value = float(schedule.iloc[0]["property_value"]) if not schedule.empty else None
            last_value = float(schedule.iloc[-1]["property_value"]) if not schedule.empty else None

            rows = [
                ("Property", inputs.meta.name if inputs.meta else ""),
                ("Property kind", inputs.meta.kind if inputs.meta else ""),
                ("Base valuation date", anchor.drawdown_date),
                ("Base value", float(anchor.property_price)),
                ("Annual growth", growth_to_decimal(anchor.property_growth_pa)),
                ("Modelling end date", inputs.modelling_end_date),
                ("Value at base date", first_value),
                ("Value at horizon", last_value),
                ("Months modelled", int(len(schedule))),
                ("LTV", "n/a (no mortgage)"),
            ]
            for k, v in rows:
                ws_s.append([k, v])

            ws_s.freeze_panes = "A2"
            ws_s["A1"].font = Font(bold=True)
            ws_s["B1"].font = Font(bold=True)
            money_keys = {"Base value", "Value at base date", "Value at horizon"}
            pct_keys = {"Annual growth"}
            date_keys = {"Base valuation date", "Modelling end date"}
            for r in range(2, ws_s.max_row + 1):
                k = ws_s.cell(row=r, column=1).value
                v = ws_s.cell(row=r, column=2)
                if k in money_keys:
                    # Currency-aware mask (euro by default) keeps a default run identical.
                    v.number_format = money_fmt
                if k in pct_keys:
                    v.number_format = "0.00%"
                if k in date_keys:
                    v.number_format = "yyyy-mm-dd"
            ws_s.column_dimensions["A"].width = 28
            ws_s.column_dimensions["B"].width = 28


def run_valuation_only(inputs: Inputs, inputs_path: Path, out_dir: Path) -> pd.DataFrame:
    """Run the no-mortgage valuation-only path and write its outputs.

    Finance note: the single entry point the CLI calls for a no-mortgage
    property. It reads the value assumptions, builds the value-over-time series,
    and writes the CSV and workbook, then prints the same "Wrote outputs to:"
    line the mortgage path prints so any wrapper script behaves consistently.
    """
    # Plain-English progress line for troubleshooting (stderr only; never stdout).
    print("[engine.valuation_only] starting valuation-only run", file=sys.stderr)

    out_dir.mkdir(parents=True, exist_ok=True)

    # Re-read the raw YAML for the dedicated valuation block: the schema loader
    # intentionally does not parse it (a no-mortgage file has no loan block).
    raw_cfg = yaml.safe_load(Path(inputs_path).read_text())

    anchor = _read_valuation_anchor(inputs, raw_cfg)
    schedule = build_valuation_schedule(inputs, anchor)
    write_valuation_outputs(out_dir, schedule, inputs, anchor)

    print("Wrote outputs to:", out_dir.resolve())
    # Plain-English completion line for troubleshooting (stderr only).
    print("[engine.valuation_only] valuation-only run complete", file=sys.stderr)
    return schedule


print("[engine.valuation_only] no-mortgage valuation-only path ready", file=sys.stderr)