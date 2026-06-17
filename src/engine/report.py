# src/engine/report.py
"""Excel formatting and portal-style summary metrics.

Finance-readable summary
------------------------
This module is the presentation and headline-number layer. It does not run the
loan simulation; it takes the already-computed schedule and event log and (a)
makes the Excel workbook readable (tables, currency/percent/date formats) and
(b) computes the "portal-style" figures a lender's online portal would show on
a chosen date, namely the principal excluding interest not yet posted and the
year-to-date interest including accrual up to that date. These feed the Summary
sheet that a reviewer reads first.

Technical summary
-----------------
Worksheet helpers ``_add_table`` and ``_format_sheet`` plus the
``compute_portal_style_metrics`` calculation. The metric reuses
``build_rate_lookup`` from ``.simulate`` to re-accrue interest day by day.

Phase 5 / S1 note: lifted verbatim out of the original ``src/engine.py`` "XLSX
helpers" and "Portal-style Summary metrics" sections. Behaviour is unchanged;
only the module header, per-function finance notes, and the stderr status line
were added.
"""

from __future__ import annotations

import sys
from datetime import date, timedelta
from typing import Dict, Optional

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .helpers import day_count_divisor, month_index
from .schema import Inputs
from .simulate import build_rate_lookup


# =====================================================================
# XLSX helpers (tables + number formats)
# =====================================================================

def _add_table(ws, name_hint="Tbl"):
    """Convert an entire worksheet into an Excel table if it has data.

    Finance note: turning each sheet into a proper Excel table is purely
    cosmetic for the reviewer (filters, striping); it does not change any
    reported figure.
    """
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 2 or max_col < 1:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    safe = "".join(ch for ch in name_hint if ch.isalnum())[:20]
    tbl = Table(displayName=f"{safe}{abs(hash((ws.title, ref)))%10000}", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def _format_sheet(ws, money_cols=None, pct_cols=None, date_cols=None, freeze=True):
    """Apply consistent formatting to a worksheet in-place.

    Finance note: this only controls how numbers look (euro, percent, dates) and
    column widths in the workbook. The underlying values are untouched, so it
    has no effect on the model's results.
    """
    money_cols = set(money_cols or [])
    pct_cols = set(pct_cols or [])
    date_cols = set(date_cols or [])

    # Header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    if freeze:
        ws.freeze_panes = "A2"

    # Header name → column index
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    def _fmt(col_name, numfmt):
        c = headers.get(col_name)
        if not c:
            return
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = numfmt

    for col in money_cols:
        _fmt(col, "€#,##0.00")
    for col in pct_cols:
        _fmt(col, "0.00%")
    for col in date_cols:
        _fmt(col, "yyyy-mm-dd")

    # Simple width heuristic
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value or ""
        ws.column_dimensions[get_column_letter(c)].width = min(max(10, len(str(header)) + 2), 28)


# =====================================================================
# Portal-style Summary metrics
# =====================================================================

def compute_portal_style_metrics(
    snapshot_date: Optional[date], inputs: Inputs, events_df: pd.DataFrame, monthly: pd.DataFrame
) -> Dict[str, float]:
    """Compute high-level portal-style metrics for summary reporting.

    Finance note: this reproduces what a borrower would see on the bank's portal
    on ``snapshot_date``: the outstanding principal excluding interest that has
    not yet been posted, and the year-to-date interest (already-posted interest
    plus accrual up to the day before the snapshot). It is the headline figure
    pair on the Summary sheet.

    ``principal_excl_unposted``
        Balance after the last posted event on or before ``snapshot_date``.

    ``ytd_interest_portal``
        Sum of posted interest in the calendar year plus the accrual for the
        current month up to (but excluding) ``snapshot_date``.  This mirrors how
        many lender portals display "accrued but not yet posted" interest.
    """
    if snapshot_date is None or events_df.empty:
        return {"principal_excl_unposted": None, "ytd_interest_portal": None}

    upto = events_df.loc[events_df["date"] <= snapshot_date].sort_values("date")
    last_bal = float(upto.iloc[-1]["balance"]) if not upto.empty else float(inputs.principal_at_drawdown)

    md = monthly.copy()
    md["posting_date"] = pd.to_datetime(md["posting_date"])
    snap_ts = pd.Timestamp(snapshot_date)

    posted_mask = (md["posting_date"].dt.year == snap_ts.year) & (md["posting_date"] <= snap_ts)
    posted_ytd = float(md.loc[posted_mask, "interest_used"].sum())

    # Accrual from day AFTER last posted interest to (snapshot_date - 1)
    last_post = (
        upto.loc[upto["kind"] == "Interest", "date"].max()
        if not upto.loc[upto["kind"] == "Interest"].empty
        else None
    )
    start = (last_post + timedelta(days=1)) if pd.notna(last_post) else inputs.drawdown_date
    end = snapshot_date - timedelta(days=1)

    accr = 0.0
    if start <= end:
        debits = (
            events_df.loc[
                (events_df["date"] >= start)
                & (events_df["date"] <= end)
                & (events_df["kind"].isin(["Payment", "Extra", "Lump"]))
            ]
            .groupby("date")["amount"]
            .sum()
            .to_dict()
        )
        before = events_df.loc[events_df["date"] < start].sort_values("date")
        bal = float(before.iloc[-1]["balance"]) if not before.empty else float(inputs.principal_at_drawdown)

        divisor = day_count_divisor(inputs.day_count)
        rate_of = build_rate_lookup(inputs.rate_blocks)
        cur = start
        while cur <= end:
            mnum = month_index(inputs.drawdown_date, cur)
            annual = rate_of(mnum)
            daily_rate = annual / divisor
            accr += bal * daily_rate
            if cur in debits:
                bal -= float(debits[cur])
            cur += timedelta(days=1)

    ytd_portal = posted_ytd + round(accr, 2)
    return {"principal_excl_unposted": last_bal, "ytd_interest_portal": ytd_portal}


print("[engine.report] xlsx helpers and portal metrics ready", file=sys.stderr)