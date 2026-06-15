# src/engine.py
"""Mortgage amortisation engine.

This module contains the full implementation of the mortgage simulator used in
the project.  The engine was written with a *single source of truth* mindset:
every step from parsing the YAML inputs to writing Excel outputs lives here so
that behaviour is easy to reason about and verify.

The code is deliberately arranged into themed sections.  Each section starts
with a banner comment so that readers can skim quickly and dive into the level
of detail they need.  The sections follow the lifecycle of the data:

1. **Helpers** – generic date and numeric helpers that have no mortgage
   awareness.
2. **Input schema** – dataclasses that describe the shape of the YAML inputs.
3. **Monthly scaffolding** – translation of daily bank events into per-month
   metadata that the simulator consumes.
4. **Engine** – the daily ACT/365 simulation that creates the canonical event
   log and monthly schedule.
5. **Reporting helpers** – Excel formatting, reconciliation, and "portal"
   metric utilities.
6. **CLI** – glue that loads inputs, runs the engine, and writes files.

Functionality has intentionally been kept identical to the previous revision;
the refactor focuses on documentation and structure so that future readers can
understand *what each block contributes to the overall goal* without having to
trace values manually.
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import yaml
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Tax module import that works both when:
# - tests import as package (from src.engine import ...), and
# - you run as script (python src\engine.py)
try:
    from .tax import load_tenancies, compute_tax_year_table  # package context
except Exception:  # pragma: no cover
    try:
        from .tax import load_tenancies, compute_tax_year_table
    except Exception:
        from tax import load_tenancies, compute_tax_year_table

# Phase 2 path resolver.  Output and tenancy locations are resolved through the
# config layer (src/paths.py) instead of being assumed relative to the current
# working directory.  The import mirrors the tax-module fallback above so the
# module works both as a package (python -m src.engine) and as a loose script.
try:
    from .paths import resolve_out_dir, resolve_relative  # package context
except Exception:  # pragma: no cover
    try:
        from .paths import resolve_out_dir, resolve_relative
    except Exception:
        from paths import resolve_out_dir, resolve_relative


# =====================================================================
# Helpers
# =====================================================================

def ensure_date(x) -> date:
    """Return a :class:`datetime.date` regardless of the input type.

    The project routinely deals with ``datetime``, ``pandas`` and YAML sourced
    values.  Normalising everything to a native ``date`` early keeps downstream
    code free from type-guarding clutter.
    """
    if isinstance(x, date):
        return x
    if isinstance(x, pd.Timestamp):
        return x.date()
    return datetime.strptime(str(x), "%Y-%m-%d").date()


def ym_int(d: date) -> int:
    """Return a ``YYYYMM`` integer that is convenient for joins/grouping."""
    return d.year * 100 + d.month


def eom(d: date) -> date:
    """Return the calendar month-end for ``d``."""
    nm = d.replace(day=28) + timedelta(days=4)
    return nm - timedelta(days=nm.day)


def month_index(d0: date, d: date) -> int:
    """Return the 1-based number of months between ``d0`` and ``d``."""
    return (d.year - d0.year) * 12 + (d.month - d0.month) + 1


def clamp_day(y: int, m: int, day: int) -> date:
    """Clamp ``(y, m, day)`` to that month's last valid day.

    This protects against invalid dates such as 31 February when we need to
    project payment dates into future months.
    """
    last = eom(date(y, m, 1)).day
    return date(y, m, min(day, last))


def pmt(rate_m: float, n_months: int, pv: float) -> float:
    """Excel-like PMT (end-of-period) returning a positive payment amount."""
    if n_months <= 0:
        return abs(pv)
    if abs(rate_m) < 1e-12:
        return pv / n_months
    r = rate_m
    return pv * (r / (1 - (1 + r) ** (-n_months)))


def day_count_divisor(label: str) -> float:
    """Return the divisor implied by the day-count convention string."""
    lab = (label or "ACT/365").upper()
    if lab in {"ACT/360", "30/360"}:
        return 360.0
    return 365.0

# ---------------------------------------------------------------------
# Property value & LTV helpers
# ---------------------------------------------------------------------

def _growth_to_decimal(g: float) -> float:
    """Normalise growth inputs to a decimal per annum.

    The project expects users to occasionally express growth in whole
    percentages (``5`` meaning 5%) while other times providing decimals.
    Returning a decimal keeps downstream math unambiguous.
    """
    try:
        g = float(g or 0.0)
    except Exception:
        g = 0.0
    return (g / 100.0) if g > 1.0 else g


def _months_between(d0: date, d1: date) -> int:
    return (d1.year - d0.year) * 12 + (d1.month - d0.month)

def property_value_on(inputs: "Inputs", dt: date) -> float:
    """Return the modelled property valuation on ``dt``.

    The valuation logic supports "revaluation blocks" where a user pins a new
    base value and growth rate from a particular date.  The implicit block at
    drawdown captures the original purchase price so that the behaviour is
    consistent whether or not custom blocks are provided.
    """
    base_price = float(inputs.property_price or 0.0)
    if base_price <= 0.0:
        return 0.0

    eff_blocks: List[ValuationBlock] = []
    imp_growth = _growth_to_decimal(getattr(inputs, "property_growth_pa", 0.0))
    eff_blocks.append(ValuationBlock(start=inputs.drawdown_date, base_value=base_price, growth_pa=imp_growth))

    if inputs.valuation_blocks:
        eff_blocks.extend(inputs.valuation_blocks)
        eff_blocks = sorted(eff_blocks, key=lambda b: b.start)

    active = None
    for b in eff_blocks:
        if b.start <= dt:
            active = b
        else:
            break

    if active is None:
        return base_price

    months = _months_between(active.start, dt)
    return float(active.base_value * ((1.0 + active.growth_pa) ** (months / 12.0)))


# =====================================================================
# Input schema
# =====================================================================

@dataclass
class RateBlock:
    """Continuous rate assumption for a span of model months."""

    start_month: int       # 1-based from drawdown
    end_month: int
    annual_rate: float     # decimal p.a., e.g. 0.0365
    kind: str              # 'fixed' or 'variable' (informational)

@dataclass
class ValuationBlock:
    """A user-specified revaluation of the property."""

    start: date          # effective date of new base valuation
    base_value: float    # the revalued amount from which growth applies
    growth_pa: float     # decimal p.a. (0.01 => 1%)

@dataclass
class Inputs:
    """Canonical representation of the YAML modelling configuration."""

    property_price: float
    principal_at_drawdown: float
    drawdown_date: date
    total_term_months: int
    first_payment_date: date
    known_first_payment: float
    repayment_day_default: int
    property_growth_pa: float           # decimal (0.01 → 1% p.a.). If >1, treated as %
    overpayment_cap_pct: float
    rate_blocks: List[RateBlock]
    strategy_at_refix: str              # 'RecalculatePayment' | 'TermReduction'
    overpay_rules: List[dict]           # standing extras (by start month)
    lump_sums: List[dict]               # exact-date one-offs: {date, amount}
    modelling_end_date: Optional[date]
    day_count: str = "ACT/365"
    # Phase 2: how to treat recurring extras when there *is* a bank payment line that month.
    # "true"  -> assume extra included in the bank Payment (suppress separate Extra)
    # "false" -> always post a separate Extra
    # "auto"  -> behave like "true" (default)
    merge_extra_mode: str = "auto"
    valuation_blocks: List[ValuationBlock] = field(default_factory=list)  # optional; overrides simple growth if provided
    reconcile_ok_abs_eur: float = 0.01
    posting_order: str = "debit_then_post"  # 'debit_then_post' | 'post_then_debit'


def load_inputs(path: Path) -> Inputs:
    """Parse the YAML modelling configuration into an :class:`Inputs` object."""
    raw = yaml.safe_load(Path(path).read_text())
    loan = raw["loan"]
    blocks = [RateBlock(**rb) for rb in raw["rate_blocks"]]
    strat = raw.get("strategy_at_refix", "RecalculatePayment")
    overp = raw.get("overpay_rules", [])
    lumps = raw.get("lump_sums", [])
    mod = raw.get("modelling", {})
    end_date = mod.get("end_date", None)

    # Phase 2 — read bank.merge_standing_extra_into_payment
    bank_cfg = (raw.get("bank") or {})
    merge_mode_raw = str(bank_cfg.get("merge_standing_extra_into_payment", "auto")).strip().lower()
    if merge_mode_raw in {"1", "true", "yes"}:
        merge_mode = "true"
    elif merge_mode_raw in {"0", "false", "no"}:
        merge_mode = "false"
    else:
        merge_mode = "auto"

    # Optional property valuation blocks (re/valuations + growth regime changes)
    vblocks_raw = (loan.get("valuation_blocks") or [])
    vblocks: List[ValuationBlock] = []

    def _to_dec(g):
        try:
            g = float(g or 0.0)
        except Exception:
            g = 0.0
        return (g / 100.0) if g > 1.0 else g

    for vb in vblocks_raw:
        vblocks.append(
            ValuationBlock(
                start=ensure_date(vb["start"]),
                base_value=float(vb["value"]),
                growth_pa=_to_dec(vb.get("growth_pa", loan.get("property_growth_pa", 0.0))),
            )
        )
    vblocks = sorted(vblocks, key=lambda b: b.start)

    # Reconcile config (absolute EUR tolerance)
    rec_cfg = (raw.get("reconcile") or {})
    try:
        ok_abs = float(rec_cfg.get("ok_abs_eur", 0.01))
    except Exception:
        ok_abs = 0.01

    # Bank posting order
    post_ord = str(bank_cfg.get("posting_order", "debit_then_post")).strip().lower()
    if post_ord not in {"post_then_debit", "debit_then_post"}:
        post_ord = "debit_then_post"

    return Inputs(
        property_price=float(loan["property_price"]),
        principal_at_drawdown=float(loan["principal_at_drawdown"]),
        drawdown_date=ensure_date(loan["drawdown_date"]),
        total_term_months=int(loan["total_term_months"]),
        first_payment_date=ensure_date(loan["first_payment_date"]),
        known_first_payment=float(loan["known_first_payment"]),
        repayment_day_default=int(loan["repayment_day_default"]),
        property_growth_pa=float(loan.get("property_growth_pa", 0.0)),
        overpayment_cap_pct=float(loan.get("overpayment_cap_pct", 0.10)),
        rate_blocks=blocks,
        strategy_at_refix=str(strat),
        overpay_rules=overp,
        lump_sums=lumps,
        modelling_end_date=(ensure_date(end_date) if end_date else None),
        day_count=str(mod.get("day_count", "ACT/365")),
        merge_extra_mode=merge_mode,
        valuation_blocks=vblocks,
        reconcile_ok_abs_eur=ok_abs,
        posting_order=post_ord,
    )


def load_actuals(csv_path: Path) -> pd.DataFrame:
    """Load bank statement events from the CSV exported by the lender.

    The helper keeps the transformation logic in one place so that tests and
    CLI invocations agree on how to interpret the CSV.  Amount sign conventions
    match the bank feed: payments are negative while interest and drawdown
    amounts are positive.
    """
    df = pd.read_csv(csv_path, parse_dates=["date"])
    df["date"] = df["date"].dt.date
    df["ym"] = df["date"].apply(ym_int)
    df["type"] = df["type"].astype(str).str.strip().str.title()
    if "run_balance" not in df.columns:
        df["run_balance"] = np.nan
    return df


# =====================================================================
# Monthly scaffolding
# =====================================================================

def build_rate_lookup(blocks: List[RateBlock]):
    """Return a callable mapping model month numbers to annual rates."""
    def rate_of_month(m: int) -> float:
        for rb in blocks:
            if rb.start_month <= m <= rb.end_month:
                return rb.annual_rate
        return blocks[-1].annual_rate
    return rate_of_month


def derive_modelling_end(inputs: Inputs) -> date:
    """Return the last date to simulate based on inputs and optional overrides."""
    if inputs.modelling_end_date:
        return inputs.modelling_end_date
    y = inputs.drawdown_date.year + (inputs.drawdown_date.month - 1 + inputs.total_term_months) // 12
    m = (inputs.drawdown_date.month - 1 + inputs.total_term_months) % 12 + 1
    return date(y, m, 5)


def month_span(start: date, end: date) -> List[date]:
    """Return a list of first-of-month dates between ``start`` and ``end``."""
    out: List[date] = []
    cur = date(start.year, start.month, 1)
    lim = date(end.year, end.month, 1)
    while cur <= lim:
        out.append(cur)
        cur = date(cur.year + (cur.month // 12), (cur.month % 12) + 1, 1)
    return out


def month_tables(inputs: Inputs, actuals: pd.DataFrame) -> pd.DataFrame:
    """Build a per-month summary table consumed by the daily engine.

    The table distils raw bank activity into the bare minimum metadata the
    simulation needs.  Deriving this table up-front keeps the inner daily loop
    simple and fast.
    """
    end = derive_modelling_end(inputs)
    mstarts = month_span(inputs.drawdown_date, end)
    g = actuals.groupby("ym", dropna=False)

    # Expand standing extras into {month_num: amount}
    recurring_by_month: Dict[int, float] = {}
    for rule in inputs.overpay_rules:
        s = int(rule["start_month"])
        amt = float(rule["amount"])
        endm = rule.get("end_month", None)
        rep = str(rule.get("repeat", "monthly")).lower()
        if rep == "monthly":
            i = s
            while i <= inputs.total_term_months and (endm is None or i <= int(endm)):
                recurring_by_month[i] = recurring_by_month.get(i, 0.0) + amt
                i += 1
        else:
            recurring_by_month[s] = recurring_by_month.get(s, 0.0) + amt

    rows = []
    for ms in mstarts:
        ym = ym_int(ms)
        mnum = month_index(inputs.drawdown_date, ms)
        eom_d = eom(ms)
        grp = g.get_group(ym) if ym in g.groups else None

        pay_date = None
        pay_total = 0.0
        post_date = None
        post_amt = 0.0

        if grp is not None:
            pays = grp[grp["type"] == "Payment"]
            if not pays.empty:
                pay_date = ensure_date(min(pays["date"]))
                pay_total = float(-pays["amount"].sum())  # input has payments negative
            ints = grp[grp["type"] == "Interest"]
            if not ints.empty:
                post_date = ensure_date(min(ints["date"]))
                post_amt = float(ints["amount"].sum())

        # Default payment date only if no actual payment that month
        def_pay = None
        if ms >= date(inputs.first_payment_date.year, inputs.first_payment_date.month, 1):
            def_pay = clamp_day(ms.year, ms.month, inputs.repayment_day_default)

        rows.append(
            dict(
                month_start=ms,
                eom=eom_d,
                ym=ym,
                month_num=mnum,
                actual_payment_date=pay_date,
                actual_payment_total=pay_total,
                actual_interest_post_date=post_date,
                actual_interest_amount=post_amt,
                default_payment_date=def_pay,
                recurring_extra=recurring_by_month.get(mnum, 0.0),
            )
        )
    return pd.DataFrame(rows)


# =====================================================================
# Engine (daily simulation)
# =====================================================================

def payment_for_month(
    inputs: Inputs, mnum: int, annual_rate: float, balance: float, months: pd.DataFrame, last_known_payment_base: float
) -> float:
    """Return the scheduled payment for a model month when none is provided.

    The logic mimics how lenders adjust standing orders:

    * First payment month uses the known value from the inputs file.
    * At the beginning of a new rate block, and only when the strategy is
      ``RecalculatePayment``, we recompute a PMT using the remaining term and
      current balance.
    * Otherwise the last known *base* payment (excluding recurring extras) is
      carried forward.
    """
    if mnum == month_index(inputs.drawdown_date, inputs.first_payment_date):
        return inputs.known_first_payment

    is_block_start = any((mnum == rb.start_month) for rb in inputs.rate_blocks)
    if inputs.strategy_at_refix == "RecalculatePayment" and is_block_start:
        remaining = max(0, inputs.total_term_months - (mnum - 1))
        return pmt(annual_rate / 12.0, remaining, balance)

    return last_known_payment_base


def run_engine(inputs: Inputs, actuals: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Run the core daily simulation.

    The function returns three DataFrames:

    ``monthly``
        Per-month schedule used for reporting and tax computations.

    ``reconcile``
        A join between modelled and bank events that supports audits and
        tolerance checks.

    ``events_df``
        The canonical event log with one row per bank-like transaction
        generated by the engine.

    Behaviour intentionally mimics observed bank behaviour.  When a set of
    debits would push the principal below zero we walk the day's transactions
    backwards trimming amounts so that the closing balance is exactly zero and
    no further interest is posted that day.
    """
    rate_of = build_rate_lookup(inputs.rate_blocks)
    months = month_tables(inputs, actuals)

    # ------------------------------------------------------------------
    # Precompute month lookups to avoid per-day DataFrame slicing.  The
    # simulation loop is tight and runs for potentially thousands of days, so
    # we prefer constant-time lookups over repeated groupby operations.
    # ------------------------------------------------------------------
    months = months.copy()
    months["pay_date"] = months.apply(
        lambda r: r.actual_payment_date if pd.notna(r.actual_payment_date) else r.default_payment_date,
        axis=1
    )
    months["has_actual_payment"] = months["actual_payment_total"] > 0
    months_by_ym = {int(r.ym): r for _, r in months.iterrows()}

    # One-off lump-sum overpayments keyed by exact date for O(1) lookups.
    lumps_by_date = {ensure_date(x["date"]): float(x["amount"]) for x in inputs.lump_sums} if inputs.lump_sums else {}

    # Rolling state of the simulation.
    start = inputs.drawdown_date
    end = derive_modelling_end(inputs)
    balance = float(inputs.principal_at_drawdown)
    accrued = 0.0

    # Base scheduled payment tracker (without standing extras).
    last_known_payment_base = float(inputs.known_first_payment)

    # Merge behaviour for recurring extras versus payments.  Users can ask the
    # model to explicitly separate recurring extras even if the bank merges
    # them into the payment line.
    merge_mode = (inputs.merge_extra_mode or "auto").strip().lower()
    # In actual-payment months: 'auto' behaves like 'true'
    actuals_merge_extras = (merge_mode in {"auto", "true"})
    # In scheduled months: only 'true' merges extras into the Payment
    scheduled_merge_extras = (merge_mode == "true")

    # Per-month collector dictionaries – keyed by ``YYYYMM`` integers.
    month_interest_used: Dict[int, float] = {int(r.ym): 0.0 for _, r in months.iterrows()}
    month_paid: Dict[int, float] = {int(r.ym): 0.0 for _, r in months.iterrows()}
    month_extras: Dict[int, float] = {int(r.ym): 0.0 for _, r in months.iterrows()}
    month_lumps: Dict[int, float] = {int(r.ym): 0.0 for _, r in months.iterrows()}
    month_rate: Dict[int, float] = {int(r.ym): rate_of(int(r.month_num)) for _, r in months.iterrows()}

    # Event log (only days where something happens).
    events: List[Dict] = []

    # Daily walk ----------------------------------------------------------------
    cur = start
    while cur <= end and balance > 0.01:
        ym = ym_int(cur)
        mrow = months_by_ym[ym]
        mnum = int(mrow.month_num)
        annual = month_rate[ym]

        # Accrual for the day ----------------------------------------------------
        daily_rate = annual / day_count_divisor(inputs.day_count)
        accrued += balance * daily_rate

        # Snapshot start-of-day principal (excl. unposted interest).
        start_balance_today = balance
        day_events_start_idx = len(events)
        loan_cleared_today = False

        # Values used regardless of whether today is a payment day or not.
        recurring_amt = float(mrow.recurring_extra or 0.0)
        lump_today = float(lumps_by_date.get(cur, 0.0))
        payment_today = 0.0
        extra_today = 0.0
        base_sched: Optional[float] = None

        is_payment_day = (pd.notna(mrow.pay_date) and ensure_date(mrow.pay_date) == cur)
        has_actual_payment = bool(mrow.has_actual_payment)

        if is_payment_day and has_actual_payment:
            # ACTUAL month: the bank decided the debit; optionally un-merge extras.
            payment_today = float(mrow.actual_payment_total)
            if not actuals_merge_extras and recurring_amt > 0.0:
                # Split bank debit into base + extra, capped by the actual amount
                base = max(0.0, payment_today - recurring_amt)
                extra_today = min(recurring_amt, payment_today)
                payment_today = base
            else:
                extra_today = 0.0

        elif is_payment_day:
            # SCHEDULED month: compute base PMT (no bank line).
            base_sched = payment_for_month(inputs, mnum, annual, balance, months, last_known_payment_base)
            if scheduled_merge_extras and recurring_amt > 0.0:
                payment_today = base_sched + recurring_amt
                extra_today = 0.0
            else:
                payment_today = base_sched
                extra_today = recurring_amt

        else:
            # Not a payment day; only lumps may occur.
            pass

        # Carry forward ONLY if we computed a new base scheduled payment.
        if base_sched is not None:
            last_known_payment_base = base_sched

        # Logic to calculate potential interest posting amount
        interest_to_post = 0.0
        should_post_interest = False
        post_date = mrow.actual_interest_post_date if pd.notna(mrow.actual_interest_post_date) else eom(cur)
        post_date = ensure_date(post_date)
        
        if cur == post_date:
            should_post_interest = True
            if (mrow.actual_interest_amount or 0.0) > 0:
                interest_to_post = float(mrow.actual_interest_amount)
            else:
                interest_to_post = round(accrued, 2)

        # ---------------- PRE-DEBIT POSTING ----------------
        if inputs.posting_order == "post_then_debit" and should_post_interest:
            balance += interest_to_post
            accrued = 0.0
            month_interest_used[ym] = interest_to_post
            events.append(dict(date=cur, kind="Interest", amount=interest_to_post, balance=round(balance, 2)))
            should_post_interest = False  # Consumed

        # ---------------- DEBITS ----------------
        # Apply debits in the order Payment → Extra → Lump to mirror the intent
        # of the inputs and make the later "trim" logic deterministic.
        # STRICT ROUNDING: All debits must be 2 decimal places before hitting the balance.
        payment_today = round(payment_today, 2)
        extra_today = round(extra_today, 2)
        lump_today = round(lump_today, 2)

        if payment_today > 0:
            balance -= payment_today
            month_paid[ym] += payment_today
            events.append(dict(date=cur, kind="Payment", amount=payment_today, balance=round(balance, 2)))

        if extra_today > 0:
            balance -= extra_today
            month_extras[ym] += extra_today
            events.append(dict(date=cur, kind="Extra", amount=extra_today, balance=round(balance, 2)))

        if lump_today > 0:
            balance -= lump_today
            month_lumps[ym] += lump_today
            events.append(dict(date=cur, kind="Lump", amount=lump_today, balance=round(balance, 2)))

        # ---- Final-payment trim (bank behaviour) ----
        if balance < -1e-9:
            overshoot = -balance
            # walk today's events backwards: Lump -> Extra -> Payment (reverse of application)
            # CAUTION: If we posted interest first, we DO NOT trim the interest event.
            # We only trim user debits (Payment/Extra/Lump).
            i = len(events) - 1
            while overshoot > 1e-9 and i >= day_events_start_idx:
                ev = events[i]
                if ev["kind"] in {"Payment", "Extra", "Lump"} and ev["amount"] > 0:
                    take = min(overshoot, float(ev["amount"]))
                    new_amt = round(float(ev["amount"]) - take, 2)
                    if ev["kind"] == "Payment":
                        month_paid[ym] -= take
                    elif ev["kind"] == "Extra":
                        month_extras[ym] -= take
                    else:
                        month_lumps[ym] -= take
                    overshoot = round(overshoot - take, 2)

                    if new_amt <= 0.0 + 1e-9:
                        events.pop(i)
                    else:
                        ev["amount"] = new_amt
                i -= 1

            # Recompute balances for today's remaining events from start_balance_today
            # (Note: start_balance_today does not include the interest if it was posted today.
            #  So we must account for that if post_then_debit happened)
            
            # Simple re-run balance calculation for stability:
            b_recalc = start_balance_today
            for j in range(day_events_start_idx, len(events)):
                e = events[j]
                op = 1.0 if e["kind"] == "Interest" else -1.0
                b_recalc += (float(e["amount"]) * op)
                e["balance"] = round(b_recalc, 2)

            balance = max(0.0, b_recalc) # Should be 0.0
            loan_cleared_today = True
            accrued = 0.0  # suppress interest posting later today (if debit-then-post)

        # ---------------- POST-DEBIT POSTING ----------------
        # Interest posting on bank date if present; else EOM.  The posting is
        # suppressed if the loan was cleared earlier today to mimic the bank's
        # behaviour.
        if inputs.posting_order == "debit_then_post" and should_post_interest and not loan_cleared_today:
            balance += interest_to_post
            accrued = 0.0
            month_interest_used[ym] = interest_to_post
            events.append(dict(date=cur, kind="Interest", amount=interest_to_post, balance=round(balance, 2)))

        # Next day --------------------------------------------------------------
        cur += timedelta(days=1)

    # ------------------ Build outputs ------------------

    # Event log DataFrame — stable ordering (Payment, Extra, Lump, Interest).
    if events:
        order = {"Payment": 0, "Extra": 1, "Lump": 2, "Interest": 3}
        events_df = pd.DataFrame(events)
        events_df["__order"] = events_df["kind"].map(order).fillna(99).astype(int)
        events_df = events_df.sort_values(["date", "__order"]).drop(columns="__order")
    else:
        events_df = pd.DataFrame(columns=["date", "kind", "amount", "balance"])

    # Per-event property value & LTV after each event
    if not events_df.empty:
        try:
            events_df["property_value"] = events_df["date"].apply(lambda d: property_value_on(inputs, ensure_date(d)))
            events_df["ltv_after_event"] = events_df["balance"] / events_df["property_value"]
        except Exception:
            events_df["property_value"] = np.nan
            events_df["ltv_after_event"] = np.nan

    # Monthly schedule ---------------------------------------------------------
    rows = []
    for _, r in months.iterrows():
        ymkey = int(r.ym)
        pay = month_paid[ymkey]; extra = month_extras[ymkey]; lump = month_lumps[ymkey]
        interest_used = month_interest_used[ymkey]
        principal = max(0.0, (pay + extra + lump) - interest_used)
        rows.append(dict(
            ym=ymkey,
            month_start=r.month_start,
            payment_date=r.pay_date,
            payment_amount=round(pay, 2),
            extra_amount=round(extra, 2),
            lump_amount=round(lump, 2),
            interest_used=round(interest_used, 2),
            principal_paid=round(principal, 2),
            annual_rate=month_rate[ymkey],
            bank_posted_interest_present=(float(r.actual_interest_amount or 0.0) > 0.0)
        ))
    monthly = pd.DataFrame(rows)

    # Posting date/year --------------------------------------------------------
    monthly["posting_date"] = [
        r.actual_interest_post_date if pd.notna(r.actual_interest_post_date) else r.eom
        for _, r in months.iterrows()
    ]
    monthly["posting_year"] = pd.to_datetime(monthly["posting_date"]).dt.year

    # Add model month-end balance from events (last event in that month).
    if not events_df.empty:
        events_df["ym"] = events_df["date"].apply(ym_int)
        eom_bal = events_df.groupby("ym", as_index=True)["balance"].last().rename("model_eom_balance")
        monthly = monthly.merge(eom_bal, left_on="ym", right_index=True, how="left")

    # Add bank month-end running balance (last bank line in that month, if provided).
    if "run_balance" in actuals.columns:
        bank_month_end = (
            actuals
            .assign(ym=lambda d: d["date"].apply(ym_int))
            .sort_values("date")
            .groupby("ym", as_index=True)["run_balance"]
            .last()
            .rename("bank_eom_running_balance")
        )
        monthly = monthly.merge(bank_month_end, left_on="ym", right_index=True, how="left")
        if "model_eom_balance" in monthly.columns:
            monthly["eom_diff_model_minus_bank"] = monthly["model_eom_balance"] - monthly["bank_eom_running_balance"]

    # Property value (at EOM) and LTVs ----------------------------------------
    try:
        monthly["property_value"] = monthly["month_start"].apply(lambda ms: property_value_on(inputs, eom(ensure_date(ms))))
        if "model_eom_balance" in monthly.columns:
            monthly["ltv_model_eom"] = monthly["model_eom_balance"] / monthly["property_value"]
        if "bank_eom_running_balance" in monthly.columns:
            monthly["ltv_bank_eom"] = monthly["bank_eom_running_balance"] / monthly["property_value"]
    except Exception:
        monthly["property_value"] = np.nan
        monthly["ltv_model_eom"] = np.nan
        if "bank_eom_running_balance" in monthly.columns:
            monthly["ltv_bank_eom"] = np.nan

    # Reconcile: join by date + type (avoids mismatches on same date).
    model_ev = events_df[events_df["kind"].isin(["Payment", "Interest"])].copy()
    model_ev.rename(columns={
        "date": "bank_date",
        "kind": "type",
        "amount": "model_amount",
        "balance": "model_balance",
    }, inplace=True)

    bank_ev = actuals[actuals["type"].isin(["Payment", "Interest"])].copy()
    bank_ev.rename(columns={"date": "bank_date", "run_balance": "bank_running_balance"}, inplace=True)

    rec = pd.merge(
        bank_ev,
        model_ev[["bank_date", "type", "model_amount", "model_balance"]],
        on=["bank_date", "type"], how="left"
    )

    # Coerce numerics and compute diffs/tolerance if both sides exist.
    for c in ("bank_running_balance", "model_balance"):
        if c in rec.columns:
            rec[c] = pd.to_numeric(rec[c], errors="coerce")

    if {"bank_running_balance", "model_balance"}.issubset(set(rec.columns)):
        rec["diff_model_minus_bank"] = rec["model_balance"] - rec["bank_running_balance"]

        # Legacy 1c label (kept for backwards compatibility)
        rec["ok_within_1c"] = rec["diff_model_minus_bank"].abs().le(0.01).map({True: "OK", False: "CHECK"})

        # NEW: configurable absolute-EUR threshold
        thr = float(getattr(inputs, "reconcile_ok_abs_eur", 0.01) or 0.01)
        rec["ok_within_abs_eur"] = rec["diff_model_minus_bank"].abs().le(thr)
        rec["ok_label"] = rec["ok_within_abs_eur"].map({True: "OK", False: "CHECK"})
        rec["ok_reason"] = rec.apply(
            lambda r: (f"|diff|≤€{thr:,.2f}" if pd.notna(r.get("diff_model_minus_bank"))
                       and abs(float(r["diff_model_minus_bank"])) <= thr else ">threshold"),
            axis=1
        )

    return monthly, rec, events_df


# =====================================================================
# XLSX helpers (tables + number formats)
# =====================================================================

def _add_table(ws, name_hint="Tbl"):
    """Convert an entire worksheet into an Excel table if it has data."""
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 2 or max_col < 1:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    safe = "".join(ch for ch in name_hint if ch.isalnum())[:20]
    tbl = Table(displayName=f"{safe}{abs(hash((ws.title, ref)))%10000}", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def _format_sheet(ws, money_cols=None, pct_cols=None, date_cols=None, freeze=True):
    """Apply consistent formatting to a worksheet in-place."""
    money_cols = set(money_cols or [])
    pct_cols = set(pct_cols or [])
    date_cols = set(date_cols or [])

    # Header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    if freeze:
        ws.freeze_panes = "A2"

    # Header name → column index
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    def _fmt(col_name, numfmt):
        c = headers.get(col_name)
        if not c:
            return
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).number_format = numfmt

    for col in money_cols:
        _fmt(col, "€#,##0.00")
    for col in pct_cols:
        _fmt(col, "0.00%")
    for col in date_cols:
        _fmt(col, "yyyy-mm-dd")

    # Simple width heuristic
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value or ""
        ws.column_dimensions[get_column_letter(c)].width = min(max(10, len(str(header)) + 2), 28)


# =====================================================================
# Portal-style Summary metrics
# =====================================================================

def compute_portal_style_metrics(
    snapshot_date: Optional[date], inputs: Inputs, events_df: pd.DataFrame, monthly: pd.DataFrame
) -> Dict[str, float]:
    """Compute high-level portal-style metrics for summary reporting.

    ``principal_excl_unposted``
        Balance after the last posted event on or before ``snapshot_date``.

    ``ytd_interest_portal``
        Sum of posted interest in the calendar year plus the accrual for the
        current month up to (but excluding) ``snapshot_date``.  This mirrors how
        many lender portals display "accrued but not yet posted" interest.
    """
    if snapshot_date is None or events_df.empty:
        return {"principal_excl_unposted": None, "ytd_interest_portal": None}

    upto = events_df.loc[events_df["date"] <= snapshot_date].sort_values("date")
    last_bal = float(upto.iloc[-1]["balance"]) if not upto.empty else float(inputs.principal_at_drawdown)

    md = monthly.copy()
    md["posting_date"] = pd.to_datetime(md["posting_date"])
    snap_ts = pd.Timestamp(snapshot_date)

    posted_mask = (md["posting_date"].dt.year == snap_ts.year) & (md["posting_date"] <= snap_ts)
    posted_ytd = float(md.loc[posted_mask, "interest_used"].sum())

    # Accrual from day AFTER last posted interest to (snapshot_date - 1)
    last_post = (
        upto.loc[upto["kind"] == "Interest", "date"].max()
        if not upto.loc[upto["kind"] == "Interest"].empty
        else None
    )
    start = (last_post + timedelta(days=1)) if pd.notna(last_post) else inputs.drawdown_date
    end = snapshot_date - timedelta(days=1)

    accr = 0.0
    if start <= end:
        debits = (
            events_df.loc[
                (events_df["date"] >= start)
                & (events_df["date"] <= end)
                & (events_df["kind"].isin(["Payment", "Extra", "Lump"]))
            ]
            .groupby("date")["amount"]
            .sum()
            .to_dict()
        )
        before = events_df.loc[events_df["date"] < start].sort_values("date")
        bal = float(before.iloc[-1]["balance"]) if not before.empty else float(inputs.principal_at_drawdown)

        divisor = day_count_divisor(inputs.day_count)
        rate_of = build_rate_lookup(inputs.rate_blocks)
        cur = start
        while cur <= end:
            mnum = month_index(inputs.drawdown_date, cur)
            annual = rate_of(mnum)
            daily_rate = annual / divisor
            accr += bal * daily_rate
            if cur in debits:
                bal -= float(debits[cur])
            cur += timedelta(days=1)

    ytd_portal = posted_ytd + round(accr, 2)
    return {"principal_excl_unposted": last_bal, "ytd_interest_portal": ytd_portal}


# =====================================================================
# CLI — write CSV + formatted XLSX (values-only)
# =====================================================================

def main():
    """Entry point used by ``python -m src.engine`` and ``src/engine.py``.

    The CLI is intentionally lightweight: load configuration, run the engine,
    and write CSV/XLSX artefacts.  Any consumer that needs to re-use the logic
    should import :func:`run_engine` directly instead of shelling out.
    """

    ap = argparse.ArgumentParser(description="Daily mortgage engine (ACT/365)")
    ap.add_argument("--inputs", type=Path, required=True, help="Path to inputs.yaml")
    ap.add_argument("--actuals", type=Path, required=True, help="Path to actuals.csv")
    # --out is optional now.  When omitted, the output folder is resolved through
    # the Phase 2 config layer (CLI > MORTGAGE_OUT_DIR > paths.local.yaml > <repo>/out).
    ap.add_argument("--out", type=Path, default=None, help="Output folder (overrides config)")
    args = ap.parse_args()

    # Resolve the output directory through the config layer.  Passing the raw CLI
    # value (or None) keeps an explicit --out as the highest-priority source.
    out_dir = resolve_out_dir(str(args.out) if args.out is not None else None)
    out_dir.mkdir(parents=True, exist_ok=True)

    inputs = load_inputs(args.inputs)
    actuals = load_actuals(args.actuals)

    # Core engine run ---------------------------------------------------------
    monthly, reconcile, events = run_engine(inputs, actuals)

    # ---- Derive "as-of" and quick summary stats for Summary sheet
    rec_non_na = (
        reconcile.dropna(subset=["model_balance"]) if "model_balance" in reconcile.columns else reconcile.copy()
    )
    asof_date = rec_non_na["bank_date"].max() if not rec_non_na.empty else None

    bank_bal = None
    model_bal_same = None
    diff_bal = None
    if asof_date is not None and "bank_running_balance" in rec_non_na.columns:
        bank_row = rec_non_na.loc[rec_non_na["bank_date"] == asof_date]
        if not bank_row.empty:
            bank_bal = float(bank_row["bank_running_balance"].iloc[-1])
            model_bal_same = float(bank_row["model_balance"].iloc[-1]) if "model_balance" in bank_row.columns else None
            if model_bal_same is not None:
                diff_bal = model_bal_same - bank_bal

    # Optional: prefer a newer portal snapshot if present
    raw_cfg = yaml.safe_load(Path(args.inputs).read_text())
    snaps = (raw_cfg.get("reconcile") or {}).get("snapshots") or []
    if snaps:
        latest = max(snaps, key=lambda s: ensure_date(s["date"]))
        snap_dt = ensure_date(latest["date"])
        if asof_date is None or snap_dt > asof_date:
            asof_date = snap_dt
            bank_bal = float(latest["balance"])
            portal_metrics = compute_portal_style_metrics(asof_date, inputs, events, monthly)
            model_bal_same = portal_metrics.get("principal_excl_unposted")
            diff_bal = (model_bal_same - bank_bal) if (model_bal_same is not None) else None

    # YTD (calendar-year posting)
    ytd_interest = None
    ytd_principal = None
    if asof_date is not None and "posting_date" in monthly.columns:
        md = monthly.copy()
        md["posting_date"] = pd.to_datetime(md["posting_date"])
        cond = (md["posting_date"].dt.year == pd.Timestamp(asof_date).year) & (
            md["posting_date"] <= pd.Timestamp(asof_date)
        )
        ytd_interest = float(md.loc[cond, "interest_used"].sum())
        ytd_principal = float(md.loc[cond, "principal_paid"].sum())

    # Next scheduled payment (from event log)
    next_pay_date = None
    next_pay_amt = None
    if asof_date is not None:
        future_pays = events[(events["kind"] == "Payment") & (events["date"] > asof_date)]
        if not future_pays.empty:
            row = future_pays.sort_values("date").iloc[0]
            next_pay_date = row["date"]
            next_pay_amt = float(row["amount"])

    # Current annual rate at as-of
    cur_rate = None
    if asof_date is not None and "ym" in monthly.columns and "annual_rate" in monthly.columns:
        ym_key = int(pd.Timestamp(asof_date).year * 100 + pd.Timestamp(asof_date).month)
        hit = monthly.loc[monthly["ym"] == ym_key]
        if not hit.empty:
            cur_rate = float(hit.iloc[0]["annual_rate"])

    # Property value path (coerce 1.00 ↦ 1% if user wrote percent)
    prop_val = None
    ltv = None
    try:
        growth = float(inputs.property_growth_pa or 0.0)
        if growth > 1.0:
            growth = growth / 100.0
        if asof_date is not None:
            m_since = (pd.Timestamp(asof_date).year - inputs.drawdown_date.year) * 12 + (
                pd.Timestamp(asof_date).month - inputs.drawdown_date.month
            )
            prop_val = float(inputs.property_price * ((1.0 + growth) ** (m_since / 12.0)))
            if bank_bal is not None and prop_val > 0:
                ltv = bank_bal / prop_val
    except Exception:
        pass

    # ---- Tax (optional) ----
    tax_cfg = (raw_cfg.get("tax") or {})
    tax_enabled = bool(tax_cfg.get("enabled", False))

    tax_year_df = None
    ten_log_df = None
    tax_audit_df = None
    tenancies = None
    policy = None

    if tax_enabled:
        # Tax computations are optional and live in ``src.tax``.  They reuse the
        # monthly schedule produced earlier and therefore inherit the same
        # assumptions as the engine.
        # Resolve tenancy files relative to the inputs.yaml folder so they live
        # beside each property's config instead of a cwd-relative data/ folder.
        # resolve_relative leaves absolute paths untouched, so explicit absolute
        # overrides still work.
        pref = resolve_relative(args.inputs, tax_cfg.get("tenancy_file", "tenancy.local.yaml"))
        fb = resolve_relative(args.inputs, "tenancy.sample.yaml")
        tenancies, policy, _ = load_tenancies(pref, fb)
        tax_year_df, ten_log_df = compute_tax_year_table(monthly, raw_cfg, tenancies, policy)

        # monthly audit only when tax is enabled
        if (tax_cfg.get("audit") or {}).get("write_monthly", True):
            from src.tax import compute_tax_monthly_audit
            tax_audit_df = compute_tax_monthly_audit(monthly, raw_cfg, tenancies, policy)

    # ---- Write outputs ----
    # XLSX
    with pd.ExcelWriter(out_dir / "mortgage_outputs.xlsx", engine="openpyxl") as xl:
        monthly.to_excel(xl, sheet_name="Monthly", index=False)
        reconcile.to_excel(xl, sheet_name="Reconcile", index=False)
        events.to_excel(xl, sheet_name="EventsDaily", index=False)

        wb = xl.book
        ws_m = xl.sheets["Monthly"]
        ws_r = xl.sheets["Reconcile"]
        ws_e = xl.sheets["EventsDaily"]

        # Tables + formats
        _add_table(ws_m, "Monthly")
        _add_table(ws_r, "Reconcile")
        _add_table(ws_e, "EventsDaily")

        _format_sheet(
            ws_m,
            money_cols=[
                "payment_amount", "extra_amount", "lump_amount",
                "interest_used", "principal_paid",
                "model_eom_balance", "bank_eom_running_balance", "eom_diff_model_minus_bank",
                "property_value"
            ],
            pct_cols=["annual_rate", "ltv_model_eom", "ltv_bank_eom"],
            date_cols=["month_start", "payment_date", "posting_date"]
        )

        _format_sheet(
            ws_r,
            money_cols=["amount", "bank_running_balance", "model_amount", "model_balance", "diff_model_minus_bank"],
            date_cols=["bank_date"],
        )
        
        _format_sheet(
            ws_e,
            money_cols=["amount", "balance", "property_value"],
            pct_cols=["ltv_after_event"],
            date_cols=["date"]
        )

        # ---------------- Tax sheets (optional) ----------------
        if tax_enabled and tax_year_df is not None:
            tax_year_df.to_excel(xl, sheet_name="TaxYear", index=False)
            ws_t = xl.sheets["TaxYear"]
            _add_table(ws_t, "TaxYear")
            _format_sheet(
                ws_t,
                money_cols=["interest_posted", "allowable_interest_s97", "principal_paid"],
                pct_cols=["avg_occupancy_ratio", "deductible_pct"],
                date_cols=[],
            )

            ten_log_df.to_excel(xl, sheet_name="TenancyLog", index=False)
            ws_log = xl.sheets["TenancyLog"]
            _add_table(ws_log, "TenancyLog")
            _format_sheet(
                ws_log,
                money_cols=["rent_amount", "security_deposit"],
                pct_cols=[],
                date_cols=["start", "end", "rtb_registration_date"],
            )
        if tax_enabled and tax_audit_df is not None:
            tax_audit_df.to_excel(xl, sheet_name="TaxAudit", index=False)
            ws_a = xl.sheets["TaxAudit"]
            _add_table(ws_a, "TaxAudit")
            _format_sheet(
                ws_a,
                money_cols=["interest_used", "principal_paid", "allowable_interest_s97"],
                pct_cols=["occupancy_ratio", "deductible_pct"],
                date_cols=["month_start", "posting_date"]
            )

        # ---------------- Summary (values only) ----------------
        ws_s = wb.create_sheet("Summary")
        ws_s.append(["Metric", "Value"])

        portal = (
            compute_portal_style_metrics(asof_date, inputs, events, monthly)
            if asof_date
            else {"principal_excl_unposted": None, "ytd_interest_portal": None}
        )

        rows = [
            ("As of date (latest bank actual)", asof_date),
            ("Bank running balance (as-of)", bank_bal),
            ("Model balance same date", model_bal_same),
            ("Difference (model - bank)", diff_bal),
            ("Current annual rate", cur_rate),
            ("Next payment date", next_pay_date),
            ("Next payment amount", next_pay_amt),
            ("YTD interest (posted only)", ytd_interest),
            ("YTD principal (posted only)", ytd_principal),
            ("Portal-style principal (excl. unposted interest)", portal["principal_excl_unposted"]),
            ("Portal YTD interest (posted + accrual to yesterday)", portal["ytd_interest_portal"]),
            ("Property value (as-of)", prop_val),
            ("LTV (as-of)", ltv),
        ]
        for k, v in rows:
            ws_s.append([k, v])

        ws_s.freeze_panes = "A2"
        ws_s["A1"].font = Font(bold=True)
        ws_s["B1"].font = Font(bold=True)
        money_keys = {
            "Bank running balance (as-of)",
            "Model balance same date",
            "Difference (model - bank)",
            "Next payment amount",
            "Property value (as-of)",
            "Portal-style principal (excl. unposted interest)",
            "Portal YTD interest (posted + accrual to yesterday)",
        }
        pct_keys = {"LTV (as-of)", "Current annual rate"}
        date_keys = {"As of date (latest bank actual)", "Next payment date"}
        for r in range(2, ws_s.max_row + 1):
            k = ws_s.cell(row=r, column=1).value
            v = ws_s.cell(row=r, column=2)
            if k in money_keys:
                v.number_format = "€#,##0.00"
            if k in pct_keys:
                v.number_format = "0.00%"
            if k in date_keys:
                v.number_format = "yyyy-mm-dd"
        ws_s.column_dimensions["A"].width = 48
        ws_s.column_dimensions["B"].width = 28

    # CSVs
    monthly.to_csv(out_dir / "schedule_monthly.csv", index=False)
    reconcile.to_csv(out_dir / "reconcile.csv", index=False)
    events.to_csv(out_dir / "events_daily.csv", index=False)
    if tax_enabled and tax_year_df is not None:
        tax_year_df.to_csv(out_dir / "tax_year.csv", index=False)
    if tax_enabled and tax_audit_df is not None:
        tax_audit_df.to_csv(out_dir / "tax_audit.csv", index=False)

    print("Wrote outputs to:", out_dir.resolve())


if __name__ == "__main__":
    main()